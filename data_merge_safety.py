# -*- coding: utf-8 -*-
"""Non-destructive recovery/merge layer for CloudDB/PostgreSQL.

Never replaces operational data. Imported workbooks/backups are merged into the
current database. Existing current rows always win; only blank fields are filled.
ID collisions between different records are remapped to a LEGACY-* identifier.
"""
import hashlib
import io
import json
import os
import re
import uuid
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
import psycopg

TZ = ZoneInfo("Asia/Riyadh")


def _s(v):
    return str(v or "").strip()


def _stable_id(prefix, payload):
    h = hashlib.sha1(payload.encode("utf-8", "ignore")).hexdigest()[:20].upper()
    return f"{prefix}-{h}"


def _int(v, default=0):
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _read_sheet(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = [_s(c.value) for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or v == "" for v in row):
            continue
        out.append({headers[i]: (row[i] if i < len(row) and row[i] is not None else "") for i in range(len(headers))})
    return out


def _parse_source(source_path):
    source_path = os.path.abspath(source_path)
    if not os.path.isfile(source_path):
        raise ValueError("ملف الاستيراد غير موجود")
    lower = source_path.lower()
    images = {}
    if lower.endswith(".zip"):
        with open(source_path, "rb") as f:
            blob = f.read()
        with zipfile.ZipFile(io.BytesIO(blob), "r") as zf:
            for info in zf.infolist():
                name = info.filename.replace("\\", "/")
                if name.startswith("/") or ".." in name.split("/"):
                    raise ValueError("ملف الاستيراد يحتوي مسارًا غير آمن")
            names = {n.replace("\\", "/"): n for n in zf.namelist()}
            db_name = "data/pharmacy_orders.xlsx" if "data/pharmacy_orders.xlsx" in names else "pharmacy_orders.xlsx"
            if db_name not in names:
                raise ValueError("لم يتم العثور على ملف pharmacy_orders.xlsx داخل النسخة")
            xlsx = zf.read(names[db_name])
            images = {k[8:]: zf.read(v) for k, v in names.items() if k.startswith("uploads/") and not k.endswith("/")}
    elif lower.endswith(".xlsx"):
        with open(source_path, "rb") as f:
            xlsx = f.read()
    else:
        raise ValueError("اختر ملف Excel (.xlsx) أو نسخة احتياطية (.zip)")
    return xlsx, images


def _merge_workbook(self, xlsx_bytes, images=None):
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=False)
    if "Orders" not in wb.sheetnames:
        wb.close()
        raise ValueError("ملف البيانات لا يحتوي ورقة Orders المطلوبة")
    orders = _read_sheet(wb, "Orders")
    if not any(_s(r.get("Order_ID")) or _s(r.get("Customer_Name")) or _s(r.get("Phone")) for r in orders):
        wb.close()
        raise ValueError("ملف البيانات لا يبدو كملف طلبات صالح")
    items = _read_sheet(wb, "Order_Items")
    logs = _read_sheet(wb, "Activity_Log")
    undos = _read_sheet(wb, "Undo_History")
    settings = _read_sheet(wb, "Settings")
    wb.close()

    created_orders = created_items = created_logs = created_undos = created_images = 0
    remapped_orders = {}
    remapped_items = {}

    # A point-in-time backup of the CURRENT state happens before any merge.
    backup_name = self.create_manual_backup(reason="auto")

    with self._connect() as conn:
        current_orders = {}
        for row in conn.execute("SELECT * FROM orders").fetchall():
            d = dict(row)
            current_orders[_s(d.get("order_id"))] = d

        # Orders: current state always wins. Legacy only fills missing values.
        for r in orders:
            old_id = _s(r.get("Order_ID"))
            if not old_id:
                continue
            if old_id in current_orders:
                cur = current_orders[old_id]
                legacy_identity = (_s(r.get("Customer_Name")), _s(r.get("Phone")), _s(r.get("Order_Date")))
                current_identity = (_s(cur.get("customer_name")), _s(cur.get("phone")), _s(cur.get("order_date")))
                compatible = (not legacy_identity[0] or not current_identity[0] or legacy_identity[0] == current_identity[0]) and (not legacy_identity[1] or not current_identity[1] or legacy_identity[1] == current_identity[1]) and (not legacy_identity[2] or not current_identity[2] or legacy_identity[2] == current_identity[2])
                if compatible:
                    fields = {
                        "customer_name": r.get("Customer_Name"), "phone": r.get("Phone"), "product_name": r.get("Product_Name"),
                        "quantity": _int(r.get("Quantity"), 0), "order_date": r.get("Order_Date"), "available_date": r.get("Available_Date"),
                        "status": r.get("Status"), "contact_status": r.get("Contact_Status"), "last_contact_date": r.get("Last_Contact_Date"),
                        "next_followup_date": r.get("Next_Followup_Date"), "pickup_date": r.get("Pickup_Date"), "notes": r.get("Notes"),
                    }
                    sets, vals = [], []
                    for col, value in fields.items():
                        if (_s(cur.get(col)) == "") and _s(value) != "":
                            sets.append(f"{col}=%s"); vals.append(_s(value) if col != "quantity" else _int(value))
                    if sets:
                        sets.append("updated_at=%s"); vals.append(datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")); vals.append(old_id)
                        conn.execute(f"UPDATE orders SET {', '.join(sets)} WHERE order_id=%s", vals)
                else:
                    new_id = f"LEGACY-{old_id}-{uuid.uuid4().hex[:8].upper()}"
                    remapped_orders[old_id] = new_id
                    oid = new_id
            else:
                oid = old_id
            if oid == old_id and old_id in current_orders:
                continue
            status = _s(r.get("Status")) or "بانتظار التوفر"
            contact = _s(r.get("Contact_Status")) or ("بانتظار رد العميل" if status in ("تم التواصل - بانتظار الاستلام", "لم يستلم") else "لم يتم التواصل")
            now = datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
            conn.execute("INSERT INTO orders(order_id,customer_name,phone,product_name,quantity,order_date,available_date,status,contact_status,last_contact_date,next_followup_date,pickup_date,notes,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                         (oid, _s(r.get("Customer_Name")), _s(r.get("Phone")), _s(r.get("Product_Name")), _int(r.get("Quantity")), _s(r.get("Order_Date")), _s(r.get("Available_Date")), status, contact, _s(r.get("Last_Contact_Date")), _s(r.get("Next_Followup_Date")), _s(r.get("Pickup_Date")), _s(r.get("Notes")), _s(r.get("Created_At")) or now, _s(r.get("Updated_At")) or now))
            current_orders[oid] = {"order_id": oid}
            created_orders += 1

        # Build the set of current item IDs and merge legacy items.
        current_items = {}
        for row in conn.execute("SELECT * FROM order_items").fetchall():
            current_items[_s(row["item_id"])] = dict(row)
        for r in items:
            old_iid = _s(r.get("Item_ID"))
            old_oid = _s(r.get("Order_ID"))
            if not old_iid or not old_oid:
                continue
            oid = remapped_orders.get(old_oid, old_oid)
            if oid != old_oid and old_iid:
                iid = f"LEGACY-{old_iid}-{uuid.uuid4().hex[:8].upper()}"
                remapped_items[old_iid] = iid
            elif old_iid in current_items:
                cur = current_items[old_iid]
                if _s(cur.get("order_id")) != oid:
                    iid = f"LEGACY-{old_iid}-{uuid.uuid4().hex[:8].upper()}"
                    remapped_items[old_iid] = iid
                else:
                    iid = old_iid
                    fields = {"image_path": r.get("Image_Path"), "availability_status": r.get("Availability_Status"), "available_price": r.get("Available_Price"), "discounted_price": r.get("Discounted_Price"), "unavailable_reason": r.get("Unavailable_Reason"), "availability_note": r.get("Availability_Note"), "price_confirmation_required": r.get("Price_Confirmation_Required"), "available_at": r.get("Available_At")}
                    sets, vals = [], []
                    for col, value in fields.items():
                        if _s(cur.get(col)) == "" and _s(value) != "":
                            sets.append(f"{col}=%s"); vals.append(_s(value))
                    if sets:
                        vals.append(iid); conn.execute(f"UPDATE order_items SET {', '.join(sets)} WHERE item_id=%s", vals)
                    continue
            else:
                iid = old_iid
            if conn.execute("SELECT 1 FROM orders WHERE order_id=%s", (oid,)).fetchone() is None:
                continue
            conn.execute("INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                         (iid, oid, _s(r.get("Product_Name")), _int(r.get("Quantity"), 1), _s(r.get("Image_Path")), _s(r.get("Availability_Status")) or "بانتظار التوفر", _s(r.get("Available_Price")), _s(r.get("Discounted_Price")), _s(r.get("Unavailable_Reason")), _s(r.get("Availability_Note")), _s(r.get("Price_Confirmation_Required")), _s(r.get("Available_At")), _s(r.get("Created_At")) or datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")))
            current_items[iid] = {"item_id": iid, "order_id": oid}
            created_items += 1

        # Logs are append-only. Use deterministic IDs to avoid duplicate imports.
        for r in logs:
            raw_order = _s(r.get("Order_ID")); oid = remapped_orders.get(raw_order, raw_order)
            payload = "|".join(_s(r.get(k)) for k in ("Log_ID","Order_ID","Action","Old_Status","New_Status","Note","Created_At","User"))
            log_id = _stable_id("IMPLOG", payload)
            if conn.execute("SELECT 1 FROM activity_log WHERE log_id=%s", (log_id,)).fetchone():
                continue
            conn.execute("INSERT INTO activity_log(log_id,order_id,action,old_status,new_status,note,created_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                         (log_id, oid, _s(r.get("Action")), _s(r.get("Old_Status")), _s(r.get("New_Status")), _s(r.get("Note")), _s(r.get("Created_At")) or datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S"), _s(r.get("User")) or "موظف"))
            created_logs += 1

        # Undo history is historical metadata; import it append-only with unique IDs.
        for r in undos:
            raw_order = _s(r.get("Order_ID")); oid = remapped_orders.get(raw_order, raw_order)
            payload = "|".join(_s(r.get(k)) for k in ("Undo_ID","Order_ID","Action","Snapshot_JSON","Created_At","Undone_At","User"))
            undo_id = _stable_id("IMPUNDO", payload)
            if conn.execute("SELECT 1 FROM undo_history WHERE undo_id=%s", (undo_id,)).fetchone():
                continue
            snapshot = _s(r.get("Snapshot_JSON")) or "{}"
            if raw_order and raw_order != oid:
                snapshot = snapshot.replace(raw_order, oid)
            conn.execute("INSERT INTO undo_history(undo_id,order_id,action,snapshot_json,created_at,undone_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                         (undo_id, oid, _s(r.get("Action")), snapshot, _s(r.get("Created_At")) or datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S"), _s(r.get("Undone_At")), _s(r.get("User")) or "موظف"))
            created_undos += 1

        # Settings are preserved from CURRENT state. Add only missing keys.
        for r in settings:
            key = _s(r.get("Key")); value = _s(r.get("Value"))
            if not key:
                continue
            if conn.execute("SELECT 1 FROM settings WHERE key=%s", (key,)).fetchone() is None:
                conn.execute("INSERT INTO settings(key,value) VALUES (%s,%s)", (key, value))

        # Images: never overwrite an existing path; map legacy item/order IDs when needed.
        if images:
            for legacy_path, data in images.items():
                path = legacy_path.replace("\\", "/").lstrip("/")
                parts = path.split("/", 1)
                if len(parts) != 2:
                    continue
                raw_oid, filename = parts
                oid = remapped_orders.get(raw_oid, raw_oid)
                raw_iid = filename.split("_", 1)[0]
                iid = remapped_items.get(raw_iid, raw_iid)
                if not conn.execute("SELECT 1 FROM orders WHERE order_id=%s", (oid,)).fetchone():
                    continue
                if not conn.execute("SELECT 1 FROM order_items WHERE item_id=%s AND order_id=%s", (iid, oid)).fetchone():
                    continue
                ext = os.path.splitext(filename)[1].lower()
                ctype = {".jpg":"image/jpeg", ".jpeg":"image/jpeg", ".png":"image/png", ".webp":"image/webp"}.get(ext, "application/octet-stream")
                target = path
                if conn.execute("SELECT 1 FROM item_images WHERE image_path=%s", (target,)).fetchone():
                    stem, ext2 = os.path.splitext(target)
                    target = f"{stem}_legacy_{uuid.uuid4().hex[:8]}{ext2}"
                conn.execute("INSERT INTO item_images(image_path,order_id,item_id,filename,content_type,data,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)", (target, oid, iid, os.path.basename(filename), ctype, psycopg.Binary(data), datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")))
                if not _s((conn.execute("SELECT image_path FROM order_items WHERE item_id=%s", (iid,)).fetchone() or {}).get("image_path")):
                    conn.execute("UPDATE order_items SET image_path=%s WHERE item_id=%s", (target, iid))
                created_images += 1

    return {
        "success": True,
        "backup": backup_name,
        "created_orders": created_orders,
        "created_items": created_items,
        "created_logs": created_logs,
        "created_undos": created_undos,
        "created_images": created_images,
        "remapped_orders": len(remapped_orders),
        "remapped_items": len(remapped_items),
        "order_count": len(self.get_all_orders()),
    }


def _install_undo_patch(self):
    """Fix the PostgreSQL undo duplicate-updated_at bug without changing behavior."""
    original_undo = self.undo_last
    if getattr(original_undo, "_ezz_undo_fixed", False):
        return

    def fixed(order_id, user="موظف"):
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM undo_history WHERE order_id=%s AND COALESCE(undone_at,'')='' ORDER BY created_at DESC LIMIT 1", (str(order_id),)).fetchone()
            current = self._fetch_order(conn, order_id)
            if not row:
                return {'error': 'لا يوجد إجراء يمكن التراجع عنه لهذا الطلب', 'code': 409}
            if not current:
                return {'error': 'الطلب غير موجود', 'code': 404}
            snapshot = json.loads(row['snapshot_json'])
            order_data = snapshot.get('order', {})
            fields = [('Customer_Name','customer_name'),('Phone','phone'),('Product_Name','product_name'),('Quantity','quantity'),('Order_Date','order_date'),('Available_Date','available_date'),('Status','status'),('Contact_Status','contact_status'),('Last_Contact_Date','last_contact_date'),('Next_Followup_Date','next_followup_date'),('Pickup_Date','pickup_date'),('Notes','notes'),('Created_At','created_at')]
            sets = [f'{dbkey}=%s' for _, dbkey in fields]
            params = [order_data.get(key, '') for key, _ in fields]
            sets.append('updated_at=%s'); params.append(datetime.now(TZ).strftime('%Y-%m-%d %H:%M:%S')); params.append(str(order_id))
            conn.execute(f"UPDATE orders SET {', '.join(sets)} WHERE order_id=%s", params)
            conn.execute('DELETE FROM order_items WHERE order_id=%s', (str(order_id),))
            for item in snapshot.get('items', []):
                conn.execute('INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)', tuple(item.get(k, '') for k in ('Item_ID','Order_ID','Product_Name','Quantity','Image_Path','Availability_Status','Available_Price','Discounted_Price','Unavailable_Reason','Availability_Note','Price_Confirmation_Required','Available_At','Created_At')))
            conn.execute('UPDATE undo_history SET undone_at=%s WHERE undo_id=%s', (datetime.now(TZ).strftime('%Y-%m-%d %H:%M:%S'), row['undo_id']))
            self._log(conn, order_id, f"تراجع عن: {row['action']}", current['status'], order_data.get('Status',''), 'تم التراجع عن آخر تغيير للمستخدم', user)
            return {'order': self._refresh_order_in_conn(conn, order_id), 'undone_action': row['action']}

    fixed._ezz_undo_fixed = True
    self.undo_last = fixed


def install_data_merge_safety(db):
    if db.__class__.__module__ != "cloud_db":
        return
    if getattr(db, "_ezz_data_merge_safety_v1", False):
        return
    original_import = db.import_legacy_data
    original_restore = db.restore_backup

    def safe_import(source_path):
        xlsx, images = _parse_source(source_path)
        return _merge_workbook(db, xlsx, images)

    def safe_restore(filename):
        safe = os.path.basename(str(filename or "")).strip()
        with db._connect() as conn:
            row = conn.execute("SELECT data FROM backups WHERE filename=%s", (safe,)).fetchone()
        if not row:
            return False
        xlsx, images = _parse_source_from_bytes(bytes(row['data']))
        result = _merge_workbook(db, xlsx, images)
        return bool(result and result.get('success'))

    def _parse_source_from_bytes(blob):
        with zipfile.ZipFile(io.BytesIO(blob), 'r') as zf:
            for info in zf.infolist():
                name = info.filename.replace('\\', '/')
                if name.startswith('/') or '..' in name.split('/'):
                    raise ValueError('ملف النسخة الاحتياطية يحتوي مسارًا غير آمن')
            names = {n.replace('\\', '/'): n for n in zf.namelist()}
            db_name = 'data/pharmacy_orders.xlsx' if 'data/pharmacy_orders.xlsx' in names else 'pharmacy_orders.xlsx'
            if db_name not in names:
                raise ValueError('النسخة الاحتياطية لا تحتوي ملف البيانات')
            xlsx = zf.read(names[db_name])
            images = {k[8:]: zf.read(v) for k, v in names.items() if k.startswith('uploads/') and not k.endswith('/')}
        return xlsx, images

    db.import_legacy_data = safe_import
    db.restore_backup = safe_restore
    _install_undo_patch(db)
    db._ezz_data_merge_safety_v1 = True
