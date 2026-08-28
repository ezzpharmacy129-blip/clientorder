import base64
import hashlib
import hmac
import html
import os
import time
from collections import defaultdict, deque

from flask import jsonify, redirect, request, session, url_for, send_file

AUTH_USERNAME = os.environ.get("AUTH_USERNAME", "admin")
AUTH_PASSWORD_HASH = os.environ.get("AUTH_PASSWORD_HASH", "")
ATTEMPTS = defaultdict(deque)
WINDOW = 600
MAX_ATTEMPTS = 5

# Arabic keyboard-layout equivalents for Latin keys used in login credentials.
_ARABIC_KEYBOARD_TO_LATIN = str.maketrans({
    "ذ":"`", "ض":"q", "ص":"w", "ث":"e", "ق":"r", "ف":"t", "غ":"y", "ع":"u", "ه":"i", "خ":"o", "ح":"p",
    "ج":"[", "د":"]", "ش":"a", "س":"s", "ي":"d", "ب":"f", "ل":"g", "ا":"h", "ت":"j", "ن":"k", "م":"l", "ك":";",
    "ط":"'", "ئ":"z", "ء":"x", "ؤ":"c", "ر":"v", "ى":"n", "ة":"m", "و":",", "ز":".", "ظ":"/",
    "٠":"0", "١":"1", "٢":"2", "٣":"3", "٤":"4", "٥":"5", "٦":"6", "٧":"7", "٨":"8", "٩":"9",
    "۰":"0", "۱":"1", "۲":"2", "۳":"3", "۴":"4", "۵":"5", "۶":"6", "۷":"7", "۸":"8", "۹":"9",
})

def _normalize_login_input(value):
    value = str(value or "")
    value = value.replace("لا", "b")
    return value.translate(_ARABIC_KEYBOARD_TO_LATIN)

def _verify(password):
    try:
        scheme, iterations_s, salt_s, digest_s = AUTH_PASSWORD_HASH.split("$", 3)
        if scheme != "pbkdf2_sha256": return False
        iterations = int(iterations_s)
        salt = base64.urlsafe_b64decode(salt_s + "=" * (-len(salt_s) % 4))
        expected = base64.urlsafe_b64decode(digest_s + "=" * (-len(digest_s) % 4))
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False

def _login_page(error=""):
    msg = f'<div style="background:#fff0f0;color:#a62b2b;padding:10px;border-radius:10px;margin-bottom:12px">{html.escape(error)}</div>' if error else ""
    return f'''<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>تسجيل الدخول - صيدلية عز الصحة</title><style>body{{margin:0;min-height:100vh;display:grid;place-items:center;background:#f4f7fb;font-family:Tahoma,Arial,sans-serif;color:#18324a}}.card{{width:min(420px,92vw);background:#fff;border-radius:22px;padding:30px;box-shadow:0 18px 60px rgba(20,50,80,.14);border:1px solid #e6edf4}}.logo{{width:72px;height:72px;border-radius:18px;display:block;margin:0 auto 18px;object-fit:contain}}h1{{font-size:23px;text-align:center;margin:0 0 6px}}p{{text-align:center;color:#6b7b8c;margin:0 0 24px}}label{{display:block;font-weight:700;margin:14px 0 7px}}input{{width:100%;padding:13px 14px;border:1px solid #ccd7e2;border-radius:12px;font-size:16px;outline:none}}button{{width:100%;margin-top:20px;border:0;border-radius:12px;padding:13px;background:#1e6f95;color:#fff;font-size:16px;font-weight:700;cursor:pointer}}.small{{font-size:12px;color:#94a0ac;text-align:center;margin-top:18px}}</style></head><body><div class="card"><img class="logo" src="/static/logo-mark.png" alt="شعار صيدلية عز الصحة"><h1>صيدلية عز الصحة</h1><p>تسجيل الدخول إلى نظام متابعة الطلبات</p>{msg}<form method="post" action="/login"><label>اسم المستخدم</label><input name="username" autocomplete="username" required autofocus><label>كلمة المرور</label><input name="password" type="password" autocomplete="current-password" required><button type="submit">دخول</button></form><div class="small">يمكنك الكتابة حتى لو كانت لوحة المفاتيح مضبوطة على العربية.</div></div></body></html>'''

def post_worker_init(worker):
    app = worker.wsgi
    if getattr(app, "_ezz_auth_installed", False):
        _install_export_route(app)
        return
    app._ezz_auth_installed = True
    app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SECURE=True, SESSION_COOKIE_SAMESITE="Lax", PERMANENT_SESSION_LIFETIME=12 * 60 * 60)
    secret = os.environ.get("SECRET_KEY")
    if secret: app.secret_key = secret

    @app.route("/login", methods=["GET", "POST"], endpoint="_ezz_login")
    def _ezz_login():
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown").split(",")[0].strip()
        now = time.time(); q = ATTEMPTS[ip]
        while q and now - q[0] > WINDOW: q.popleft()
        if request.method == "GET":
            if session.get("authenticated"): return redirect(url_for("index"))
            return _login_page()
        if len(q) >= MAX_ATTEMPTS: return _login_page("تم تجاوز عدد المحاولات المسموح بها. حاول لاحقًا."), 429
        q.append(now)
        username = _normalize_login_input(request.form.get("username", "").strip())
        password = _normalize_login_input(request.form.get("password", ""))
        if username == _normalize_login_input(AUTH_USERNAME) and AUTH_PASSWORD_HASH and _verify(password):
            session.clear(); session.permanent = True; session["authenticated"] = True; session["username"] = AUTH_USERNAME
            return redirect(url_for("index"))
        return _login_page("اسم المستخدم أو كلمة المرور غير صحيحة."), 401

    @app.route("/logout", methods=["GET", "POST"], endpoint="_ezz_logout")
    def _ezz_logout():
        session.clear(); return redirect(url_for("_ezz_login"))

    def _guard():
        path = request.path
        if path in ("/login", "/logout") or path.startswith("/static/"): return None
        if session.get("authenticated"): return None
        if path.startswith("/api/"): return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        return redirect(url_for("_ezz_login"))
    app.before_request_funcs.setdefault(None, []).insert(0, _guard)

    def _headers(response):
        response.headers.setdefault("Cache-Control", "no-store")
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        response.headers.setdefault("Referrer-Policy", "same-origin")
        return response
    app.after_request_funcs.setdefault(None, []).insert(0, _headers)
    _install_export_route(app)

def _install_export_route(app):
    """Register a protected, read-only Excel export endpoint on the actual Flask app."""
    if "export_current_xlsx_direct" in app.view_functions:
        return
    from io import BytesIO
    from datetime import datetime
    from zoneinfo import ZoneInfo
    from openpyxl import Workbook
    from db import db, ORDERS_HEADERS, ITEM_HEADERS, LOG_HEADERS, UNDO_HEADERS

    @app.get("/api/data/export-xlsx", endpoint="export_current_xlsx_direct")
    def export_current_xlsx_direct():
        if not session.get("authenticated") and not session.get("user_id"):
            return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        try:
            orders = db.get_all_orders()
            wb = Workbook(); ws = wb.active; ws.title = "Orders"; ws.append(ORDERS_HEADERS)
            for order in orders: ws.append([order.get(k, "") for k in ORDERS_HEADERS])
            wi = wb.create_sheet("Order_Items"); wi.append(ITEM_HEADERS)
            for order in orders:
                for item in order.get("Items", []) or []: wi.append([item.get(k, "") for k in ITEM_HEADERS])
            wl = wb.create_sheet("Activity_Log"); wl.append(LOG_HEADERS)
            try:
                for row in db.get_activity_log(): wl.append([row.get(k, "") for k in LOG_HEADERS])
            except Exception: pass
            wu = wb.create_sheet("Undo_History"); wu.append(UNDO_HEADERS)
            try:
                for row in getattr(db, "get_all_undo_history", lambda: [])(): wu.append([row.get(k, "") for k in UNDO_HEADERS])
            except Exception: pass
            for sheet in wb.worksheets: sheet.freeze_panes = "A2"
            buf = BytesIO(); wb.save(buf); wb.close(); buf.seek(0)
            stamp = datetime.now(ZoneInfo("Asia/Riyadh")).strftime("%Y-%m-%d_%H%M%S")
            return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Ezz_Pharmacy_Backup_{stamp}.xlsx")
        except Exception as exc:
            app.logger.exception("Excel export failed")
            return jsonify({"error": f"تعذر تصدير البيانات: {exc}"}), 500
