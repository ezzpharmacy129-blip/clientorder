import re
from flask import request, jsonify


def install_security_extensions(app, db):
    if getattr(app, "_ezz_security_extensions", False):
        return
    app._ezz_security_extensions = True
    auth = app.extensions["ezz_auth"]
    current_user = auth["current_user"]
    audit = auth["audit"]
    is_cloud = bool(auth.get("is_cloud"))

    def audit_rows(limit=500):
        try:
            if is_cloud:
                rows = db.get_activity_log(None)
                return rows[:max(1, int(limit))]
            # Local fallback only: the desktop Excel backend keeps its existing log.
            import os
            from openpyxl import load_workbook
            path = getattr(db, "DB_PATH", "")
            if not path or not os.path.exists(path):
                return []
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb["Activity_Log"] if "Activity_Log" in wb.sheetnames else None
            rows = []
            if ws:
                for r in ws.iter_rows(min_row=2, values_only=True):
                    values = list(r) + [""] * 8
                    rows.append({"created_at": values[6], "user_name": values[7], "action": values[2], "order_id": values[1], "note": values[5], "old_status": values[3], "new_status": values[4]})
            wb.close()
            return rows[-max(1, int(limit)):][::-1]
        except Exception:
            return []

    @app.before_request
    def admin_data_protection():
        if request.path in ("/api/data/reset", "/api/backups/restore"):
            user = current_user()
            if not user:
                return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}), 401
            if user.get("role") != "admin":
                return jsonify({"error":"غير مصرح لك بهذا الإجراء"}), 403
        return None

    @app.get("/api/admin/audit")
    def admin_audit_api():
        user = current_user()
        if not user:
            return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}), 401
        if user.get("role") != "admin":
            return jsonify({"error":"غير مصرح لك بهذا الإجراء"}), 403
        try:
            limit = min(max(int(request.args.get("limit", 500)), 1), 2000)
        except (TypeError, ValueError):
            limit = 500
        return jsonify({"logs": audit_rows(limit)})

    @app.after_request
    def action_audit(response):
        try:
            user = current_user()
            if not user:
                return response

            # Employees can use order APIs but should not receive the full audit trail.
            if user.get("role") != "admin" and request.method == "GET" and re.match(r"^/api/orders/[^/]+$", request.path):
                payload = response.get_json(silent=True)
                if isinstance(payload, dict) and "activity_log" in payload:
                    payload.pop("activity_log", None)
                    response.set_json(payload)

            # CloudDB already writes the canonical audit event inside each
            # successful mutation transaction. Do not write a second copy here.
            if is_cloud:
                return response

            # Local Excel fallback keeps the previous high-level audit behavior.
            if response.status_code >= 400:
                return response
            path = request.path
            method = request.method
            m = re.match(r"^/api/orders/([^/]+)(?:/|$)", path)
            order_id = m.group(1) if m else ""
            action = None
            if path == "/api/whatsapp/open-shortages":
                action = "WhatsApp"
            elif path.startswith("/api/whatsapp/open/") or path.startswith("/api/whatsapp/order/"):
                action = "WhatsApp"
            elif method == "POST" and path.endswith("/contact"):
                action = "Contact Customer"
            elif method == "POST" and path.endswith("/pickup"):
                action = "Pickup"
            elif method == "POST" and path.endswith("/postpone"):
                action = "Postpone"
            elif method == "POST" and path.endswith("/cancel"):
                action = "Cancel"
            elif method == "POST" and path.endswith("/undo"):
                action = "Undo"
            elif method == "POST" and path.endswith("/availability"):
                action = "Mark Available"
            elif method == "POST" and path.endswith("/available"):
                action = "Mark Available"
            elif method == "POST" and path == "/api/orders":
                action = "Create Order"
            elif method == "PUT" and m:
                action = "Edit Order"
            elif method == "DELETE" and m and "/items/" not in path:
                action = "Delete Order"
            if action:
                audit(order_id=order_id, action=action, note="تم تنفيذ العملية")
            return response
        except Exception:
            return response

    @app.post("/api/admin/users/<uid>/delete")
    def admin_delete_user(uid):
        user = current_user()
        if not user:
            return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}), 401
        if user.get("role") != "admin":
            return jsonify({"error":"غير مصرح لك بهذا الإجراء"}), 403
        if uid == user.get("user_id"):
            return jsonify({"error":"لا يمكنك حذف حسابك الحالي"}), 400
        try:
            with db._connect() as conn:
                row = conn.execute("SELECT username,name FROM users WHERE user_id=%s", (uid,)).fetchone()
                if not row:
                    return jsonify({"error":"المستخدم غير موجود"}), 404
                conn.execute("DELETE FROM users WHERE user_id=%s", (uid,))
            audit(action="Delete User", note=f"حذف المستخدم {row['username']}")
            return jsonify({"success":True})
        except Exception:
            return jsonify({"error":"تعذر حذف المستخدم"}),500
