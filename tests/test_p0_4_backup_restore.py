import base64
import io
import os
import tempfile
import unittest
import zipfile

_ci_database_url = os.environ.pop('DATABASE_URL', None)
try:
    from cloud_db import CloudDB
    from db import STATUS_PENDING, CONTACT_NOT_CONTACTED
    import psycopg
finally:
    if _ci_database_url:
        os.environ['DATABASE_URL'] = _ci_database_url

from openpyxl import load_workbook


class BackupRestoreIntegrationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not _ci_database_url:
            raise unittest.SkipTest('DATABASE_URL not set')
        cls.db = CloudDB()

    def setUp(self):
        with self.db._connect() as conn:
            conn.execute('TRUNCATE backups, activity_log, undo_item_images, undo_history, item_images, order_items, orders')
            conn.execute(
                "INSERT INTO orders(order_id,customer_name,phone,product_name,quantity,status,contact_status,created_at,updated_at) VALUES ('ORD-RESTORE','عميل الاسترجاع','0500000000','منتج الاسترجاع',2,%s,%s,'T','T')",
                (STATUS_PENDING, CONTACT_NOT_CONTACTED),
            )
            conn.execute(
                "INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,created_at) VALUES ('ITEM-RESTORE','ORD-RESTORE','منتج الاسترجاع',2,'ORD-RESTORE/ITEM-RESTORE.png','بانتظار التوفر','T')"
            )
            conn.execute(
                "INSERT INTO activity_log(log_id,order_id,action,old_status,new_status,note,created_at,user_name) VALUES ('LOG-RESTORE','ORD-RESTORE','اختبار','','','ملاحظة','T','اختبار')"
            )
            snapshot = '{"order":{"Order_ID":"ORD-RESTORE"},"items":[{"Item_ID":"ITEM-RESTORE"}]}'
            conn.execute(
                "INSERT INTO undo_history(undo_id,order_id,action,snapshot_json,created_at,undone_at,user_name) VALUES ('UNDO-RESTORE','ORD-RESTORE','اختبار',%s,'T','','اختبار')",
                (snapshot,),
            )
            conn.execute(
                "INSERT INTO undo_item_images(undo_id,item_id,image_path,filename,content_type,data,created_at) VALUES ('UNDO-RESTORE','ITEM-RESTORE','ORD-RESTORE/ITEM-RESTORE.png','ITEM-RESTORE.png','image/png',%s,'T')",
                (psycopg.Binary(b'UNDO-IMAGE-BYTES'),),
            )
            conn.execute(
                "INSERT INTO item_images(image_path,order_id,item_id,filename,content_type,data,created_at) VALUES ('ORD-RESTORE/ITEM-RESTORE.png','ORD-RESTORE','ITEM-RESTORE','ITEM-RESTORE.png','image/png',%s,'T')",
                (psycopg.Binary(b'ACTIVE-IMAGE-BYTES'),),
            )
            conn.execute("INSERT INTO settings(key,value) VALUES ('backup_test','before') ON CONFLICT(key) DO UPDATE SET value='before'")

    def _counts(self):
        with self.db._connect() as conn:
            return {
                table: int(conn.execute(f'SELECT COUNT(*) AS c FROM {table}').fetchone()['c'])
                for table in ('orders','order_items','activity_log','undo_history','undo_item_images','item_images')
            }

    def test_backup_contains_undo_item_images(self):
        filename = self.db.create_manual_backup(reason='manual')
        with self.db._connect() as conn:
            blob = conn.execute('SELECT data FROM backups WHERE filename=%s', (filename,)).fetchone()['data']
        with zipfile.ZipFile(io.BytesIO(bytes(blob)), 'r') as zf:
            self.assertIn('data/pharmacy_orders.xlsx', zf.namelist())
            self.assertIn('uploads/ORD-RESTORE/ITEM-RESTORE.png', zf.namelist())
            workbook = load_workbook(io.BytesIO(zf.read('data/pharmacy_orders.xlsx')), read_only=True, data_only=False)
            try:
                self.assertIn('Undo_Item_Images', workbook.sheetnames)
                rows = list(workbook['Undo_Item_Images'].iter_rows(values_only=True))
            finally:
                workbook.close()
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0:5], ('UNDO-RESTORE','ITEM-RESTORE','ORD-RESTORE/ITEM-RESTORE.png','ITEM-RESTORE.png','image/png'))
        self.assertEqual(base64.b64decode(rows[1][5]), b'UNDO-IMAGE-BYTES')

    def test_restore_roundtrip_restores_undo_images_and_core_data(self):
        expected = self._counts()
        filename = self.db.create_manual_backup(reason='manual')

        with self.db._connect() as conn:
            conn.execute('UPDATE orders SET customer_name=\'تغيير قبل الاسترجاع\' WHERE order_id=\'ORD-RESTORE\'')
            conn.execute('DELETE FROM undo_item_images')
            conn.execute('DELETE FROM item_images')
            conn.execute('DELETE FROM activity_log')
            conn.execute('DELETE FROM undo_history')
            conn.execute('DELETE FROM order_items')
            conn.execute('DELETE FROM orders')
            conn.execute("UPDATE settings SET value='changed' WHERE key='backup_test'")

        self.assertTrue(self.db.restore_backup(filename))
        self.assertEqual(self._counts(), expected)
        with self.db._connect() as conn:
            active = bytes(conn.execute("SELECT data FROM item_images WHERE image_path='ORD-RESTORE/ITEM-RESTORE.png'").fetchone()['data'])
            undo = bytes(conn.execute("SELECT data FROM undo_item_images WHERE undo_id='UNDO-RESTORE'").fetchone()['data'])
            setting = conn.execute("SELECT value FROM settings WHERE key='backup_test'").fetchone()['value']
            customer = conn.execute("SELECT customer_name FROM orders WHERE order_id='ORD-RESTORE'").fetchone()['customer_name']
        self.assertEqual(active, b'ACTIVE-IMAGE-BYTES')
        self.assertEqual(undo, b'UNDO-IMAGE-BYTES')
        self.assertEqual(setting, 'before')
        self.assertEqual(customer, 'عميل الاسترجاع')

    def test_failed_restore_preserves_current_data(self):
        with self.db._connect() as conn:
            conn.execute("INSERT INTO backups(filename,reason,created_at,data) VALUES ('bad.zip','manual','T',%s)", (psycopg.Binary(b'not-a-zip'),))
        before = self._counts()
        with self.assertRaises(Exception):
            self.db.restore_backup('bad.zip')
        after = self._counts()
        self.assertEqual(after, before)
        with self.db._connect() as conn:
            customer = conn.execute("SELECT customer_name FROM orders WHERE order_id='ORD-RESTORE'").fetchone()['customer_name']
        self.assertEqual(customer, 'عميل الاسترجاع')


if __name__ == '__main__':
    unittest.main()
