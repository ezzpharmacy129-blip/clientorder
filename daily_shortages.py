# -*- coding: utf-8 -*-
"""Excel-backed Daily Pharmacy Shortages storage for the offline build."""
import json
import os
import uuid
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

TZ = ZoneInfo("Asia/Riyadh")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_ROOT = os.environ.get("EZZ_PHARMACY_DATA_DIR")
if DATA_ROOT:
    DATA_ROOT = os.path.abspath(os.path.expandvars(os.path.expanduser(DATA_ROOT)))
elif os.name == "nt" and os.environ.get("APPDATA"):
    DATA_ROOT = os.path.join(os.environ["APPDATA"], "Ezz Pharmacy Fresh")
else:
    DATA_ROOT = os.path.join(os.path.expanduser("~"), ".ezz_pharmacy_fresh")

DATA_DIR = os.path.join(DATA_ROOT, "data")
os.makedirs(DATA_DIR, exist_ok=True)
BOOK_PATH = os.path.join(DATA_DIR, "pharmacy_orders.xlsx")

SHORTAGE_HEADERS = ["shortage_id", "product_name", "quantity", "note", "status", "created_at", "updated_at", "created_by", "resolved_at"]
UNDO_HEADERS = ["undo_id", "shortage_id", "action", "snapshot_json", "created_at", "undone_at", "user_name"]


def _now():
    return datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")


def _user_name(user):
    value = str(user or "").strip()
    return value or "موظف"


def _id(prefix):
    return f"{prefix}-{uuid.uuid4().hex[:12].upper()}"


def _ensure_book():
    if not os.path.exists(BOOK_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Pharmacy_Shortages"
        ws.append(SHORTAGE_HEADERS)
        wu = wb.create_sheet("Pharmacy_Shortage_Undo")
        wu.append(UNDO_HEADERS)
        wb.save(BOOK_PATH)

    wb = load_workbook(BOOK_PATH)
    if "Pharmacy_Shortages" not in wb.sheetnames:
        ws = wb.create_sheet("Pharmacy_Shortages")
        ws.append(SHORTAGE_HEADERS)
    if "Pharmacy_Shortage_Undo" not in wb.sheetnames:
        ws = wb.create_sheet("Pharmacy_Shortage_Undo")
        ws.append(UNDO_HEADERS)
    wb.save(BOOK_PATH)


def _read_rows(sheet_name, headers):
    _ensure_book()
    wb = load_workbook(BOOK_PATH, data_only=False)
    ws = wb[sheet_name]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in values):
            continue
        row = {h: (values[i] if i < len(values) and values[i] is not None else "") for i, h in enumerate(headers)}
        rows.append(row)
    return rows


def _append(sheet_name, headers, row):
    _ensure_book()
    wb = load_workbook(BOOK_PATH)
    ws = wb[sheet_name]
    ws.append([row.get(h, "") for h in headers])
    wb.save(BOOK_PATH)


def _rewrite(sheet_name, headers, rows):
    _ensure_book()
    wb = load_workbook(BOOK_PATH)
    ws = wb[sheet_name]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(BOOK_PATH)


def ensure_schema():
    _ensure_book()


def _row(row):
    if not row:
        return None
    return {
        "shortage_id": str(row.get("shortage_id", "")),
        "product_name": str(row.get("product_name", "")),
        "quantity": int(row.get("quantity") or 0),
        "note": str(row.get("note") or ""),
        "status": str(row.get("status") or "pending"),
        "created_at": str(row.get("created_at") or ""),
        "updated_at": str(row.get("updated_at") or ""),
        "created_by": str(row.get("created_by") or "موظف"),
        "resolved_at": str(row.get("resolved_at") or ""),
    }


def _all_shortage_rows():
    return _read_rows("Pharmacy_Shortages", SHORTAGE_HEADERS)


def list_shortages(include_cancelled=False):
    rows = _all_shortage_rows()
    if not include_cancelled:
        rows = [r for r in rows if str(r.get("status") or "pending") != "cancelled"]
    rows.sort(key=lambda r: (0 if str(r.get("status") or "pending") == "pending" else 1, str(r.get("created_at") or "")), reverse=False)
    rows.sort(key=lambda r: (0 if str(r.get("status") or "pending") == "pending" else 1, str(r.get("created_at") or "")), reverse=False)
    pending = [r for r in rows if str(r.get("status") or "pending") == "pending"]
    other = [r for r in rows if str(r.get("status") or "pending") != "pending"]
    return [_row(r) for r in sorted(pending, key=lambda r: str(r.get("created_at") or ""), reverse=True) + sorted(other, key=lambda r: str(r.get("created_at") or ""), reverse=True)]


def get_shortage(shortage_id):
    sid = str(shortage_id)
    for row in _all_shortage_rows():
        if str(row.get("shortage_id")) == sid:
            return _row(row)
    return None


def _replace_shortage(updated):
    rows = _all_shortage_rows()
    sid = str(updated["shortage_id"])
    found = False
    for idx, row in enumerate(rows):
        if str(row.get("shortage_id")) == sid:
            rows[idx] = updated
            found = True
            break
    if not found:
        rows.append(updated)
    _rewrite("Pharmacy_Shortages", SHORTAGE_HEADERS, rows)


def _append_undo(shortage_id, action, snapshot, user):
    _append("Pharmacy_Shortage_Undo", UNDO_HEADERS, {
        "undo_id": _id("PSU"),
        "shortage_id": str(shortage_id),
        "action": action,
        "snapshot_json": json.dumps(snapshot, ensure_ascii=False),
        "created_at": _now(),
        "undone_at": "",
        "user_name": _user_name(user),
    })


def _invalidate_undo(shortage_id):
    rows = _read_rows("Pharmacy_Shortage_Undo", UNDO_HEADERS)
    changed = False
    now = _now()
    for row in rows:
        if str(row.get("shortage_id")) == str(shortage_id) and not str(row.get("undone_at") or ""):
            row["undone_at"] = now
            changed = True
    if changed:
        _rewrite("Pharmacy_Shortage_Undo", UNDO_HEADERS, rows)


def _log(shortage_id, action, old_status, new_status, note, user):
    # The main offline db.py owns the general Activity_Log sheet.
    # To keep this module independent, shortage actions remain recorded in the
    # dedicated undo sheet and the visible shortage row itself.
    return None


def create_shortage(product_name, quantity, note="", user="موظف"):
    name = str(product_name or "").strip()
    try:
        qty = int(quantity)
    except (TypeError, ValueError):
        qty = 0
    if not name:
        raise ValueError("اسم المنتج مطلوب")
    if qty <= 0:
        raise ValueError("الكمية يجب أن تكون رقمًا صحيحًا أكبر من صفر")
    now = _now()
    row = {
        "shortage_id": _id("PS"),
        "product_name": name,
        "quantity": qty,
        "note": str(note or "").strip(),
        "status": "pending",
        "created_at": now,
        "updated_at": now,
        "created_by": _user_name(user),
        "resolved_at": "",
    }
    _append("Pharmacy_Shortages", SHORTAGE_HEADERS, row)
    return get_shortage(row["shortage_id"])


def update_shortage(shortage_id, product_name=None, quantity=None, note=None, user="موظف"):
    sid = str(shortage_id)
    current = get_shortage(sid)
    if not current:
        return None
    name = current["product_name"] if product_name is None else str(product_name).strip()
    qty = current["quantity"] if quantity is None else int(quantity)
    memo = current["note"] if note is None else str(note).strip()
    if not name:
        raise ValueError("اسم المنتج مطلوب")
    if qty <= 0:
        raise ValueError("الكمية يجب أن تكون رقمًا صحيحًا أكبر من صفر")
    _invalidate_undo(sid)
    updated = dict(current)
    updated.update({"product_name": name, "quantity": qty, "note": memo, "updated_at": _now()})
    _append_undo(sid, "تعديل نقص صيدلية", current, user)
    _replace_shortage(updated)
    return get_shortage(sid)


def set_available(shortage_id, user="موظف"):
    sid = str(shortage_id)
    current = get_shortage(sid)
    if not current:
        return None
    if current["status"] == "available":
        return current
    _invalidate_undo(sid)
    now = _now()
    _append_undo(sid, "تم توفير المنتج", current, user)
    updated = dict(current)
    updated.update({"status": "available", "resolved_at": now, "updated_at": now})
    _replace_shortage(updated)
    return get_shortage(sid)


def undo_last(shortage_id, user="موظف"):
    sid = str(shortage_id)
    rows = _read_rows("Pharmacy_Shortage_Undo", UNDO_HEADERS)
    candidates = [r for r in rows if str(r.get("shortage_id")) == sid and not str(r.get("undone_at") or "")]
    candidates.sort(key=lambda r: str(r.get("created_at") or ""), reverse=True)
    row = candidates[0] if candidates else None
    current = get_shortage(sid)
    if not row:
        return {"error": "لا يوجد إجراء يمكن التراجع عنه لهذا النقص", "code": 409}
    if not current:
        return {"error": "النقص غير موجود", "code": 404}
    snapshot = json.loads(str(row.get("snapshot_json") or "{}"))
    now = _now()
    restored = dict(snapshot)
    restored["updated_at"] = now
    _replace_shortage(restored)
    for r in rows:
        if r.get("undo_id") == row.get("undo_id"):
            r["undone_at"] = now
            break
    _rewrite("Pharmacy_Shortage_Undo", UNDO_HEADERS, rows)
    return {"shortage": get_shortage(sid), "undone_action": row.get("action", "")}


def stats():
    rows = _all_shortage_rows()
    total = len(rows)
    pending = sum(1 for r in rows if str(r.get("status") or "pending") == "pending")
    available = sum(1 for r in rows if str(r.get("status") or "pending") == "available")
    cancelled = sum(1 for r in rows if str(r.get("status") or "pending") == "cancelled")
    return {"total": total, "pending": pending, "available": available, "cancelled": cancelled}
