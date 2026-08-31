# -*- coding: utf-8 -*-
"""Admin-only, read-only export of the PostgreSQL dataset retained across rollback."""
import os
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

from flask import jsonify, send_file, session, request
from openpyxl import Workbook
import psycopg
from psycopg.rows import dict_row


def _sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = ws.dimensions


def install_postrollback_export(app):
    if "export_postrollback" in app.view_functions:
        return

    @app.get("/api/data/export-postrollback", endpoint="export_postrollback")
    def export_postrollback():
        if not session.get("authenticated") and not session.get("user_id"):
            return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        if str(session.get("username", "")).lower() not in {"admin", "administrator"} and session.get("role") not in {"admin", "administrator"}:
            return jsonify({"error": "صلاحية المدير مطلوبة"}), 403
        url = os.environ.get("DATABASE_URL", "").strip()
        if not url:
            return jsonify({"error": "اتصال PostgreSQL غير مضبوط"}), 503
        try:
            with psycopg.connect(url, row_factory=dict_row, connect_timeout=10) as conn:
                def rows(sql):
                    return conn.execute(sql).fetchall()
                orders = rows("SELECT * FROM orders ORDER BY created_at, order_id")
                items = rows("SELECT * FROM order_items ORDER BY created_at, item_id")
                logs = rows("SELECT * FROM activity_log ORDER BY created_at, log_id")
                undo = rows("SELECT * FROM undo_history ORDER BY created_at, undo_id")
                settings = rows("SELECT * FROM settings ORDER BY key")
                image_meta = rows("SELECT image_path, order_id, item_id, filename, content_type, created_at FROM item_images ORDER BY created_at, image_path")
                backup_meta = rows("SELECT filename, reason, created_at, octet_length(data) AS bytes FROM backups ORDER BY created_at, filename")
            wb = Workbook(); ws = wb.active; ws.title = "Orders"
            _sheet(ws, ["order_id","customer_name","phone","product_name","quantity","order_date","available_date","status","contact_status","last_contact_date","next_followup_date","pickup_date","notes","created_at","updated_at"], orders)
            _sheet(wb.create_sheet("Order_Items"), ["item_id","order_id","product_name","quantity","image_path","availability_status","available_price","discounted_price","unavailable_reason","availability_note","price_confirmation_required","available_at","created_at"], items)
            _sheet(wb.create_sheet("Activity_Log"), ["log_id","order_id","action","old_status","new_status","note","created_at","user_name"], logs)
            _sheet(wb.create_sheet("Undo_History"), ["undo_id","order_id","action","snapshot_json","created_at","undone_at","user_name"], undo)
            _sheet(wb.create_sheet("Settings"), ["key","value"], settings)
            _sheet(wb.create_sheet("Image_Metadata"), ["image_path","order_id","item_id","filename","content_type","created_at"], image_meta)
            _sheet(wb.create_sheet("Backup_Metadata"), ["filename","reason","created_at","bytes"], backup_meta)
            buf = BytesIO(); wb.save(buf); wb.close(); buf.seek(0)
            stamp = datetime.now(ZoneInfo("Asia/Riyadh")).strftime("%Y-%m-%d_%H%M%S")
            return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Ezz_PostRollback_PostgreSQL_{stamp}.xlsx", max_age=0, conditional=False)
        except Exception as exc:
            app.logger.exception("Post-rollback PostgreSQL export failed")
            return jsonify({"error": f"تعذر استخراج بيانات ما قبل الـRollback: {exc}"}), 500

    # VIP_CUSTOMERS_FEATURE_V1 — isolated table and API; existing order/data tables are untouched.
    import vip_customers
    vip_customers.ensure_schema()

    def vip_auth():
        if not session.get("authenticated") and not session.get("user_id"):
            return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}), 401
        return None

    @app.get("/api/vip-customers")
    def api_vip_customers():
        denied = vip_auth()
        if denied: return denied
        try:
            vip_customers.ensure_schema()
            return jsonify({"customers": vip_customers.list_customers()})
        except Exception as exc:
            app.logger.exception("VIP customers read failed")
            return jsonify({"error": f"تعذر قراءة العملاء المميزين: {exc}"}), 500

    @app.post("/api/vip-customers")
    def api_create_vip_customer():
        denied = vip_auth()
        if denied: return denied
        data = request.get_json(silent=True) or {}
        try:
            vip_customers.ensure_schema()
            customer = vip_customers.create_customer(data.get("name"), data.get("phone"), data.get("offer_product"), data.get("offer_price"))
            return jsonify({"customer": customer}), 201
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            app.logger.exception("VIP customer create failed")
            return jsonify({"error": f"تعذر إضافة العميل المميز: {exc}"}), 500

    @app.post("/api/vip-customers/<customer_id>/status")
    def api_vip_customer_status(customer_id):
        denied = vip_auth()
        if denied: return denied
        data = request.get_json(silent=True) or {}
        try:
            vip_customers.ensure_schema()
            customer = vip_customers.set_status(customer_id, str(data.get("status") or ""))
            if customer is None:
                return jsonify({"error":"العميل غير موجود"}), 404
            return jsonify({"customer": customer})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            app.logger.exception("VIP customer status update failed")
            return jsonify({"error": f"تعذر تحديث حالة العرض: {exc}"}), 500
