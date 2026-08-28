# -*- coding: utf-8 -*-
"""Administrator-only bidirectional availability state controls."""
import json
import os
import uuid
from flask import request, jsonify
from functools import wraps


def install_admin_state_controls(app, db):
    if getattr(app, "_ezz_admin_state_controls", False):
        return
    app._ezz_admin_state_controls = True

    auth = app.extensions["ezz_auth"]
    current_user = auth["current_user"]
    audit = auth["audit"]

    def admin_only(fn):
        @wraps(fn)
        def wrapped(*args, **kwargs):
            user = current_user()
            if not user:
                return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
            if user.get("role") != "admin":
                return jsonify({"error": "تغيير حالة الطلب متاح للمدير فقط"}), 403
            return fn(*args, **kwargs)
        return wrapped

    @app.post("/api/admin/orders/<order_id>/availability-state")
    @admin_only
    def admin_availability_state(order_id):
        data = request.get_json(silent=True) or {}
        target = str(data.get("status") or "").strip()
        reason = str(data.get("reason") or "تعديل يدوي من المدير").strip()
        allowed = {
            "بانتظار التوفر": "بانتظار التوفر",
            "متوفر - يحتاج اتصال": "متوفر",
            "غير متوفر - يحتاج اتصال": "غير متوفر",
        }
        if target not in allowed:
            return jsonify({"error": "الحالة غير مدعومة"}), 400

        from openpyxl import load_workbook
        from db import DB_PATH, _lock, _make_backup, _format_sheet, LOG_HEADERS, UNDO_HEADERS, STATUS_PENDING, STATUS_AVAILABLE, STATUS_UNAVAILABLE, CONTACT_NOT_CONTACTED, now_str

        with _lock:
            if not os.path.exists(DB_PATH):
                return jsonify({"error": "ملف البيانات غير موجود"}), 500
            # Safety snapshot before every manual state change.
            try:
                _make_backup("auto")
            except Exception:
                pass

            wb = load_workbook(DB_PATH)
            try:
                ws = wb["Orders"]
                wi = wb["Order_Items"]
                wl = wb["Activity_Log"]
                wu = wb["Undo_History"]

                order_headers = {str(c.value): i + 1 for i, c in enumerate(ws[1])}
                item_headers = {str(c.value): i + 1 for i, c in enumerate(wi[1])}
                oid_col = order_headers["Order_ID"]
                status_col = order_headers["Status"]
                contact_col = order_headers["Contact_Status"]
                updated_col = order_headers["Updated_At"]
                available_date_col = order_headers["Available_Date"]
                last_contact_col = order_headers["Last_Contact_Date"]
                followup_col = order_headers["Next_Followup_Date"]
                pickup_col = order_headers["Pickup_Date"]

                order_row = None
                old_status = ""
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(r, oid_col).value or "") == str(order_id):
                        order_row = r
                        old_status = str(ws.cell(r, status_col).value or "")
                        break
                if order_row is None:
                    return jsonify({"error": "الطلب غير موجود"}), 404

                item_rows = []
                for r in range(2, wi.max_row + 1):
                    if str(wi.cell(r, item_headers["Order_ID"]).value or "") == str(order_id):
                        item_rows.append(r)
                if not item_rows:
                    return jsonify({"error": "لا توجد منتجات في الطلب"}), 409

                # Save an exact lightweight snapshot in Undo_History for this manual state operation.
                snapshot = {
                    "order_status": old_status,
                    "order_contact": ws.cell(order_row, contact_col).value or "",
                    "order_available_date": ws.cell(order_row, available_date_col).value or "",
                    "order_last_contact": ws.cell(order_row, last_contact_col).value or "",
                    "order_followup": ws.cell(order_row, followup_col).value or "",
                    "order_pickup": ws.cell(order_row, pickup_col).value or "",
                    "items": [],
                }
                for r in item_rows:
                    snapshot["items"].append({
                        "row": r,
                        "availability": wi.cell(r, item_headers["Availability_Status"]).value or "",
                        "available_price": wi.cell(r, item_headers["Available_Price"]).value or "",
                        "discounted_price": wi.cell(r, item_headers["Discounted_Price"]).value or "",
                        "reason": wi.cell(r, item_headers["Unavailable_Reason"]).value or "",
                        "availability_note": wi.cell(r, item_headers["Availability_Note"]).value or "",
                        "available_at": wi.cell(r, item_headers["Available_At"]).value or "",
                    })

                new_status = allowed[target]
                if target == "بانتظار التوفر":
                    item_availability = "بانتظار التوفر"
                elif target == "متوفر - يحتاج اتصال":
                    item_availability = "متوفر"
                else:
                    item_availability = "غير متوفر"

                for r in item_rows:
                    wi.cell(r, item_headers["Availability_Status"]).value = item_availability
                    wi.cell(r, item_headers["Availability_Note"]).value = ""
                    if target == "غير متوفر - يحتاج اتصال":
                        wi.cell(r, item_headers["Unavailable_Reason"]).value = reason
                    else:
                        wi.cell(r, item_headers["Unavailable_Reason"]).value = ""
                    if target == "بانتظار التوفر":
                        wi.cell(r, item_headers["Available_Price"]).value = ""
                        wi.cell(r, item_headers["Discounted_Price"]).value = ""
                        wi.cell(r, item_headers["Available_At"]).value = ""
                    elif target == "متوفر - يحتاج اتصال":
                        wi.cell(r, item_headers["Available_At"]).value = now_str()

                ws.cell(order_row, status_col).value = new_status
                ws.cell(order_row, contact_col).value = CONTACT_NOT_CONTACTED
                ws.cell(order_row, available_date_col).value = now_str().split(" ")[0] if target != "بانتظار التوفر" else ""
                ws.cell(order_row, last_contact_col).value = ""
                ws.cell(order_row, followup_col).value = ""
                ws.cell(order_row, pickup_col).value = ""
                ws.cell(order_row, updated_col).value = now_str()

                log_id = "LOG-" + uuid.uuid4().hex[:12]
                user = current_user()
                wl.append([log_id, str(order_id), "Manual State Change", old_status, new_status, f"تعديل حالة التوفر يدويًا: {target}", now_str(), user["name"]])

                wu.append([
                    "UNDO-" + uuid.uuid4().hex[:12], str(order_id), "Manual State Change",
                    json.dumps(snapshot, ensure_ascii=False, default=str), now_str(), "", user["name"]
                ])
                _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu)
                wb.save(DB_PATH)
            finally:
                wb.close()

        try:
            audit(order_id=order_id, action="Manual State Change", old_status=old_status, new_status=new_status, note=f"تعديل يدوي للحالة إلى {target}")
        except Exception:
            pass
        return jsonify({"success": True, "order": db.get_order(order_id)})
