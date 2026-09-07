# -*- coding: utf-8 -*-
"""Excel-backed data layer for Ezz Pharmacy order follow-up system."""
import os
import json
import shutil
import threading
import uuid
import zipfile
import tempfile
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Persistent application data lives outside the program/version folder.
# This lets users unzip a new version beside the old one without losing data.
def _default_shared_root():
    configured = os.environ.get("EZZ_PHARMACY_DATA_DIR")
    if configured:
        return os.path.abspath(os.path.expandvars(os.path.expanduser(configured)))
    if os.name == "nt":
        roaming = os.environ.get("APPDATA")
        if roaming:
            return os.path.join(roaming, "Ezz Pharmacy Fresh")
    return os.path.join(os.path.expanduser("~"), ".ezz_pharmacy_fresh")

SHARED_ROOT = _default_shared_root()
LEGACY_DATA_DIR = os.path.join(BASE_DIR, "data")
LEGACY_BACKUP_DIR = os.path.join(BASE_DIR, "backups")
LEGACY_UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DATA_DIR = os.path.join(SHARED_ROOT, "data")
BACKUP_DIR = os.path.join(SHARED_ROOT, "backups")
UPLOAD_DIR = os.path.join(SHARED_ROOT, "uploads")
DB_PATH = os.path.join(DATA_DIR, "pharmacy_orders.xlsx")
TMP_PATH = os.path.join(DATA_DIR, "pharmacy_orders.tmp.xlsx")
STORAGE_MARKER = os.path.join(SHARED_ROOT, "storage.json")
TZ = ZoneInfo("Asia/Riyadh")
MAX_BACKUPS = 100
MAX_IMAGE_SIZE = 10 * 1024 * 1024
ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

ORDERS_HEADERS = [
    "Order_ID", "Customer_Name", "Phone", "Product_Name", "Quantity",
    "Order_Date", "Available_Date", "Status", "Contact_Status", "Last_Contact_Date",
    "Next_Followup_Date", "Pickup_Date", "Notes", "Created_At", "Updated_At",
]
ITEM_HEADERS = ["Item_ID", "Order_ID", "Product_Name", "Quantity", "Image_Path", "Availability_Status", "Available_Price", "Discounted_Price", "Unavailable_Reason", "Availability_Note", "Price_Confirmation_Required", "Available_At", "Created_At", "Customer_Decision"]
LOG_HEADERS = ["Log_ID", "Order_ID", "Action", "Old_Status", "New_Status", "Note", "Created_At", "User"]
UNDO_HEADERS = ["Undo_ID", "Order_ID", "Action", "Snapshot_JSON", "Created_At", "Undone_At", "User"]
SETTINGS_HEADERS = ["Key", "Value"]

STATUS_PENDING = "بانتظار التوفر"
STATUS_AVAILABLE = "متوفر - يحتاج اتصال"
STATUS_PARTIAL = "متوفر جزئيًا - يحتاج اتصال"
STATUS_UNAVAILABLE = "غير متوفر - يحتاج اتصال"
STATUS_CONTACTED = "تم التواصل - بانتظار الاستلام"
STATUS_PICKED_UP = "تم الاستلام"
STATUS_NOT_PICKED = "لم يستلم"
STATUS_CANCELLED = "ملغي"
CONTACT_NOT_CONTACTED = "لم يتم التواصل"
CONTACT_AWAITING = "بانتظار رد العميل"
CONTACT_ACCEPTED = "العميل موافق"
CONTACT_REJECTED = "العميل رفض"
CONTACT_POSTPONED = "مؤجل"
ALL_CONTACT_STATUSES = [CONTACT_NOT_CONTACTED, CONTACT_AWAITING, CONTACT_ACCEPTED, CONTACT_REJECTED, CONTACT_POSTPONED]
ALL_STATUSES = [STATUS_PENDING, STATUS_AVAILABLE, STATUS_PARTIAL, STATUS_UNAVAILABLE, STATUS_CONTACTED, STATUS_PICKED_UP, STATUS_NOT_PICKED, STATUS_CANCELLED]
CLOSED_STATUSES = {STATUS_PICKED_UP, STATUS_CANCELLED}

DEFAULT_SETTINGS = {
    "Pharmacy_Name": "صيدلية عز الصحة",
    "System_Name": "نظام متابعة طلبات العملاء",
    "Tagline": "رعاية من القلب",
    "Default_Followup_Days": "2",
    "Message_Template_Price_Confirmation": "السلام عليكم {اسم_العميل} 🌷\n\nمعك {اسم_الصيدلية}.\n\nالمنتج الذي طلبتموه أصبح متوفرًا لدينا ✅\n\n{المنتجات_المتوفرة}\n\nهل يناسبكم السعر ونقوم بتوفيره لكم؟\n\n{الشعار} 💙",
    "Message_Template_Available": "السلام عليكم {اسم_العميل} 🌷\n\nمعك {اسم_الصيدلية}.\n\nنود إبلاغك أن طلبك أصبح متوفرًا لدينا ✅\n\n{المنتجات_المتوفرة}\n\nيسعدنا خدمتك واستقبالك لاستلام الطلب.\n\n{الشعار} 💙",
    "Message_Template_Partial": "السلام عليكم {اسم_العميل} 🌷\n\nمعك {اسم_الصيدلية}.\n\nبخصوص طلبكم، أصبح جزء من الطلب متوفرًا لدينا ✅، ونعتذر عن عدم توفر بعض المنتجات حاليًا.\n\n✅ المنتجات المتوفرة:\n{المنتجات_المتوفرة}\n\n❌ المنتجات غير المتوفرة:\n{المنتجات_غير_المتوفرة}\n\nيسعدنا خدمتك، ونعتذر عن أي إزعاج.\n{الشعار} 💙",
    "Message_Template_Unavailable": "السلام عليكم {اسم_العميل} 🌷\n\nمعك {اسم_الصيدلية}.\n\nبخصوص طلبكم، نعتذر لأن المنتجات المطلوبة غير متوفرة حاليًا. 🙏\n\n❌ المنتجات غير المتوفرة:\n{المنتجات_غير_المتوفرة}\n\nنعتذر عن الإزعاج ونسعد بخدمتكم دائمًا.\n{الشعار} 💙",
    "Message_Template_Shortage": "📦 نواقص العملاء – {اسم_الصيدلية}\nالتاريخ: {التاريخ}\n\n{النواقص}\n\nفضلاً توفير الكميات أعلاه عند الإمكان.\n{الشعار} 💙",
}

_lock = threading.RLock()


def _workbook_has_orders(path):
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ok = "Orders" in wb.sheetnames and wb["Orders"].max_row > 1
        wb.close()
        return ok
    except Exception:
        return False


def _candidate_legacy_sources():
    return []

def _bootstrap_shared_storage():
    """Create the fresh app storage without importing any previous version/data."""
    os.makedirs(SHARED_ROOT, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    return False

def now_str():
    return datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")


def today_str():
    return datetime.now(TZ).strftime("%Y-%m-%d")


def add_days(date_str, days):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return (d + timedelta(days=int(days))).strftime("%Y-%m-%d")


def _sheet_to_dicts(ws, headers):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or v == "" for v in row):
            continue
        item = {h: (row[i] if i < len(row) and row[i] is not None else "") for i, h in enumerate(headers)}
        rows.append(item)
    return rows


def _format_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            widths[c.column] = max(widths.get(c.column, 0), min(len(str(c.value or "")) + 2, 45))
    for col_idx, width in widths.items():
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(12, width)


def _safe_backup_name(reason):
    ts = datetime.now(TZ).strftime("%Y-%m-%d_%H%M%S_%f")
    reason = "manual" if reason == "manual" else "auto"
    return f"backup_{reason}_{ts}.zip"


def _prune_backups():
    files = [f for f in os.listdir(BACKUP_DIR) if f.lower().endswith((".zip", ".xlsx"))]
    files.sort(key=lambda f: os.path.getmtime(os.path.join(BACKUP_DIR, f)))
    while len(files) > MAX_BACKUPS:
        old = files.pop(0)
        try:
            os.remove(os.path.join(BACKUP_DIR, old))
        except OSError:
            pass


def _make_backup(reason="auto"):
    """Create a full snapshot of Excel + uploaded images."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    if not os.path.exists(DB_PATH):
        return None
    name = _safe_backup_name(reason)
    path = os.path.join(BACKUP_DIR, name)
    temp_path = path + ".tmp"
    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.write(DB_PATH, arcname="data/pharmacy_orders.xlsx")
            if os.path.isdir(UPLOAD_DIR):
                for root, _, files in os.walk(UPLOAD_DIR):
                    for filename in files:
                        full = os.path.join(root, filename)
                        rel = os.path.relpath(full, UPLOAD_DIR)
                        zf.write(full, arcname=os.path.join("uploads", rel).replace(os.sep, "/"))
            zf.writestr("backup_info.txt", f"Ezz Pharmacy backup\nCreated: {now_str()}\nReason: {reason}\n")
        os.replace(temp_path, path)
        _prune_backups()
        return name
    except Exception:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except OSError:
            pass
        raise


def _atomic_save(wb):
    wb.save(TMP_PATH)
    try:
        wb.close()
    except Exception:
        pass
    os.replace(TMP_PATH, DB_PATH)


def ensure_db():
    _bootstrap_shared_storage()
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    if not os.path.exists(DB_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(ORDERS_HEADERS)
        wi = wb.create_sheet("Order_Items")
        wi.append(ITEM_HEADERS)
        wl = wb.create_sheet("Activity_Log")
        wl.append(LOG_HEADERS)
        wu = wb.create_sheet("Undo_History")
        wu.append(UNDO_HEADERS)
        wsset = wb.create_sheet("Settings")
        wsset.append(SETTINGS_HEADERS)
        for k, v in DEFAULT_SETTINGS.items():
            wsset.append([k, v])
        for wsx in wb.worksheets:
            _format_sheet(wsx)
        wb.save(DB_PATH)
        wb.close()
        return False

    with _lock:
        wb = load_workbook(DB_PATH)
        changed = False
        ws_orders = wb["Orders"] if "Orders" in wb.sheetnames else wb.create_sheet("Orders", 0)
        current_order_headers = [str(c.value or "") for c in ws_orders[1]]
        if current_order_headers != ORDERS_HEADERS:
            old_rows = list(ws_orders.iter_rows(min_row=2, values_only=True))
            old_map = {h: i for i, h in enumerate(current_order_headers)}
            ws_orders.delete_rows(1, ws_orders.max_row)
            ws_orders.append(ORDERS_HEADERS)
            for vals in old_rows:
                rec = {h: (vals[i] if i < len(vals) and vals[i] is not None else "") for h, i in old_map.items()}
                old_status = rec.get("Status", STATUS_PENDING) or STATUS_PENDING
                inferred_contact = rec.get("Contact_Status", "") or (CONTACT_AWAITING if old_status in (STATUS_CONTACTED, STATUS_NOT_PICKED) else CONTACT_NOT_CONTACTED)
                ws_orders.append([
                    rec.get("Order_ID", ""), rec.get("Customer_Name", ""), rec.get("Phone", ""), rec.get("Product_Name", ""),
                    rec.get("Quantity", 0) or 0, rec.get("Order_Date", ""), rec.get("Available_Date", ""), old_status, inferred_contact,
                    rec.get("Last_Contact_Date", ""), rec.get("Next_Followup_Date", ""), rec.get("Pickup_Date", ""), rec.get("Notes", ""),
                    rec.get("Created_At", "") or now_str(), rec.get("Updated_At", "") or now_str()
                ])
            changed = True
        if "Order_Items" not in wb.sheetnames:
            wi = wb.create_sheet("Order_Items")
            wi.append(ITEM_HEADERS)
            changed = True
            existing = _sheet_to_dicts(ws_orders, ORDERS_HEADERS)
            for order in existing:
                if order.get("Product_Name"):
                    wi.append([
                        f"ITEM-{wi.max_row:06d}", order.get("Order_ID"), order.get("Product_Name"),
                        order.get("Quantity") or 1, "", "بانتظار التوفر", "", "", "", "", "", "", order.get("Created_At") or now_str()
                    ])
        else:
            wi = wb["Order_Items"]
            current_headers = [str(c.value or "") for c in wi[1]]
            # Normalize/migrate Order_Items columns without losing existing data.
            if current_headers != ITEM_HEADERS:
                old_rows = list(wi.iter_rows(min_row=2, values_only=True))
                old_map = {h: i for i, h in enumerate(current_headers)}
                wi.delete_rows(1, wi.max_row)
                wi.append(ITEM_HEADERS)
                for vals in old_rows:
                    rec = {h: (vals[i] if i < len(vals) and vals[i] is not None else "") for h, i in old_map.items()}
                    wi.append([
                        rec.get("Item_ID", ""), rec.get("Order_ID", ""), rec.get("Product_Name", ""),
                        rec.get("Quantity", 1) or 1, rec.get("Image_Path", ""),
                        rec.get("Availability_Status", "بانتظار التوفر") or "بانتظار التوفر",
                        rec.get("Available_Price", ""), rec.get("Discounted_Price", ""),
                        rec.get("Unavailable_Reason", ""), rec.get("Availability_Note", ""), rec.get("Price_Confirmation_Required", ""),
                        rec.get("Available_At", ""), rec.get("Created_At", "")
                    ])
                changed = True
        if "Activity_Log" not in wb.sheetnames:
            wb.create_sheet("Activity_Log").append(LOG_HEADERS)
            changed = True
        if "Undo_History" not in wb.sheetnames:
            wb.create_sheet("Undo_History").append(UNDO_HEADERS)
            changed = True
        if "Settings" not in wb.sheetnames:
            ws = wb.create_sheet("Settings")
            ws.append(SETTINGS_HEADERS)
            for k, v in DEFAULT_SETTINGS.items():
                ws.append([k, v])
            changed = True
        else:
            existing_keys = {r[0] for r in wb["Settings"].iter_rows(min_row=2, values_only=True) if r and r[0]}
            for k, v in DEFAULT_SETTINGS.items():
                if k not in existing_keys:
                    wb["Settings"].append([k, v])
                    changed = True
        for wsx in wb.worksheets:
            _format_sheet(wsx)
        if changed:
            _atomic_save(wb)
        else:
            wb.close()
    return True


class ExcelDB:
    def __init__(self):
        existed = ensure_db()
        self._startup_backup_created = False
        if existed:
            try:
                wb = load_workbook(DB_PATH, read_only=True, data_only=True)
                has_orders = wb["Orders"].max_row > 1 if "Orders" in wb.sheetnames else False
                wb.close()
                # Do not create a default backup when the database is completely empty.
                if has_orders:
                    _make_backup("auto")
                    self._startup_backup_created = True
            except Exception:
                # Startup should not fail only because a backup could not be created.
                self._startup_backup_created = False

    def _load(self):
        return load_workbook(DB_PATH)

    def _read_items(self, wb):
        if "Order_Items" not in wb.sheetnames:
            return []
        ws = wb["Order_Items"]
        headers = _sheet_headers_compatible(ws, ITEM_HEADERS)
        return _sheet_to_dicts(ws, headers)

    def _items_for(self, items, order_id):
        return [i for i in items if str(i.get("Order_ID")) == str(order_id)]

    def _attach_items(self, orders, items):
        grouped = {}
        for item in items:
            grouped.setdefault(str(item.get("Order_ID")), []).append(item)
        for o in orders:
            its = grouped.get(str(o.get("Order_ID")), [])
            if not its and o.get("Product_Name"):
                its = [{
                    "Item_ID": "", "Order_ID": o.get("Order_ID"), "Product_Name": o.get("Product_Name"),
                    "Quantity": o.get("Quantity") or 1, "Image_Path": "", "Availability_Status": "بانتظار التوفر", "Available_Price": "", "Discounted_Price": "", "Unavailable_Reason": "", "Availability_Note": "", "Price_Confirmation_Required": "", "Available_At": "", "Created_At": o.get("Created_At", "")
                }]
            o["Items"] = its
            if its:
                o["Product_Name"] = "، ".join(f"{i.get('Product_Name','')} × {i.get('Quantity',1)}" for i in its)
                o["Quantity"] = sum(int(i.get("Quantity") or 0) for i in its)
            else:
                o["Product_Name"] = ""
                o["Quantity"] = 0
        return orders

    def search_orders_page(self, q="", status="", date_from="", date_to="", page=1, page_size=20):
        orders=self.get_all_orders(); q=str(q or "").strip().lower(); status=str(status or "").strip()
        if q:
            trans=str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹","01234567890123456789"); qp=q.translate(trans)
            def match(o):
                blob=" ".join([str(o.get("Customer_Name","")),str(o.get("Phone","")),str(o.get("Product_Name","")),str(o.get("Order_ID",""))]).lower()
                return q in blob or (qp and qp in str(o.get("Phone",""))) or any(q in str(i.get("Product_Name","")).lower() for i in o.get("Items",[]))
            orders=[o for o in orders if match(o)]
        if status: orders=[o for o in orders if o.get("Status")==status]
        if date_from: orders=[o for o in orders if str(o.get("Order_Date",""))>=date_from]
        if date_to: orders=[o for o in orders if str(o.get("Order_Date",""))<=date_to]
        orders.sort(key=lambda o:str(o.get("Created_At","")),reverse=True); total=len(orders)
        page=max(1,int(page or 1)); page_size=max(1,min(100,int(page_size or 20))); start=(page-1)*page_size
        return {"orders":orders[start:start+page_size],"count":total,"total":total,"page":page,"page_size":page_size,"pages":max(1,(total+page_size-1)//page_size)}

    def get_all_orders(self):
        with _lock:
            wb = self._load()
            orders = _sheet_to_dicts(wb["Orders"], ORDERS_HEADERS)
            items = self._read_items(wb)
            wb.close()
        return self._attach_items(orders, items)

    def get_order(self, order_id):
        orders = self.get_all_orders()
        return next((o for o in orders if str(o.get("Order_ID")) == str(order_id)), None)

    def get_activity_log_page(self, user="", action="", order_id="", q="", page=1, page_size=50):
        rows=self.get_activity_log(None)
        user=str(user or "").strip().lower(); action=str(action or "").strip()
        order_id=str(order_id or "").strip(); q=str(q or "").strip().lower()
        if user: rows=[r for r in rows if str(r.get("User","")).lower()==user]
        if action: rows=[r for r in rows if str(r.get("Action",""))==action]
        if order_id: rows=[r for r in rows if str(r.get("Order_ID",""))==order_id]
        if q: rows=[r for r in rows if q in str(r.get("Note","")).lower() or q in str(r.get("Order_ID","")).lower() or q in str(r.get("User","")).lower()]
        total=len(rows); page=max(1,int(page or 1)); page_size=max(1,min(100,int(page_size or 50))); start=(page-1)*page_size
        return {"rows":rows[start:start+page_size],"count":total,"total":total,"page":page,"page_size":page_size,"pages":max(1,(total+page_size-1)//page_size)}

    def get_activity_log(self, order_id=None):
        with _lock:
            wb = self._load()
            logs = _sheet_to_dicts(wb["Activity_Log"], LOG_HEADERS)
            wb.close()
        if order_id is not None:
            logs = [x for x in logs if str(x.get("Order_ID")) == str(order_id)]
        logs.sort(key=lambda x: str(x.get("Created_At", "")), reverse=True)
        return logs

    def get_settings(self):
        with _lock:
            wb = self._load()
            settings = {}
            for row in wb["Settings"].iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    settings[row[0]] = row[1]
            wb.close()
            return settings

    def update_settings(self, updates):
        allowed = set(DEFAULT_SETTINGS)
        clean = {str(k): str(v) for k, v in (updates or {}).items() if str(k) in allowed}
        if not clean:
            return self.get_settings()
        with _lock:
            wb = self._load()
            ws = wb["Settings"]
            row_by_key = {}
            for row in ws.iter_rows(min_row=2):
                key = row[0].value
                if key:
                    row_by_key[str(key)] = row
            for key, value in clean.items():
                row = row_by_key.get(key)
                if row:
                    row[1].value = value
                else:
                    ws.append([key, value])
            _atomic_save(wb)
        return self.get_settings()

    def _next_order_id(self, ws):
        max_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            oid = str(row[0] or "")
            if oid.startswith("ORD-"):
                try:
                    max_num = max(max_num, int(oid.split("-")[1]))
                except (ValueError, IndexError):
                    pass
        return f"ORD-{max_num + 1:05d}"

    def _next_item_id(self, ws):
        max_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            iid = str(row[0] or "")
            if iid.startswith("ITEM-"):
                try:
                    max_num = max(max_num, int(iid.split("-")[1]))
                except (ValueError, IndexError):
                    pass
        return f"ITEM-{max_num + 1:06d}"

    def _append_log(self, ws, order_id, action, old_status, new_status, note, user):
        ws.append([f"LOG-{ws.max_row:06d}", order_id, action, old_status or "", new_status or "", note or "", now_str(), user or "موظف"])

    def _find_row(self, ws, order_id):
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and str(row[0].value) == str(order_id):
                return row
        return None

    def create_order(self, customer_name, phone, products, notes="", order_date=None, user="موظف"):
        products = [{"product_name": str(p.get("product_name", "")).strip(), "quantity": int(p.get("quantity", 0))} for p in products]
        products = [p for p in products if p["product_name"] and p["quantity"] > 0]
        if not products:
            raise ValueError("يجب إضافة منتج واحد على الأقل")
        with _lock:
            _make_backup()
            wb = self._load()
            ws = wb["Orders"]
            wi = wb["Order_Items"]
            wl = wb["Activity_Log"]
            wu = wb["Undo_History"]
            oid = self._next_order_id(ws)
            ts = now_str()
            odate = order_date or today_str()
            summary = "، ".join(f"{p['product_name']} × {p['quantity']}" for p in products)
            total_qty = sum(p["quantity"] for p in products)
            row = {
                "Order_ID": oid, "Customer_Name": customer_name, "Phone": phone,
                "Product_Name": summary, "Quantity": total_qty, "Order_Date": odate,
                "Available_Date": "", "Status": STATUS_PENDING, "Contact_Status": CONTACT_NOT_CONTACTED, "Last_Contact_Date": "",
                "Next_Followup_Date": "", "Pickup_Date": "", "Notes": notes or "",
                "Created_At": ts, "Updated_At": ts,
            }
            ws.append([row[h] for h in ORDERS_HEADERS])
            item_rows = []
            for p in products:
                item_id = self._next_item_id(wi)
                wi.append([item_id, oid, p["product_name"], p["quantity"], "", "بانتظار التوفر", "", "", "", "", "", "", ts])
                item_rows.append({"Item_ID": item_id, "Order_ID": oid, "Product_Name": p["product_name"], "Quantity": p["quantity"], "Image_Path": "", "Created_At": ts})
            self._append_log(wl, oid, "إنشاء الطلب", "", STATUS_PENDING, f"تم تسجيل {len(products)} منتج في الطلب", user)
            _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu)
            _atomic_save(wb)
            try:
                _make_backup("auto")
            except Exception:
                pass
        return self.get_order(oid)

    def _update_fields(self, ws, order_id, fields):
        row = self._find_row(ws, order_id)
        if row is None:
            return False
        idx = {h: i for i, h in enumerate(ORDERS_HEADERS)}
        for key, val in fields.items():
            if key in idx:
                row[idx[key]].value = val
        row[idx["Updated_At"]].value = now_str()
        return True

    def _status(self, ws, order_id):
        row = self._find_row(ws, order_id)
        return row[ORDERS_HEADERS.index("Status")].value if row else None

    def update_order(self, order_id, fields, products=None, user="موظف"):
        with _lock:
            _make_backup()
            wb = self._load(); ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            old = self._status(ws, order_id)
            self._invalidate_undo(wu, order_id)
            if old is None:
                wb.close(); return None
            if products is not None:
                products = [{"product_name": str(p.get("product_name", "")).strip(), "quantity": int(p.get("quantity", 0))} for p in products]
                products = [p for p in products if p["product_name"] and p["quantity"] > 0]
                if not products:
                    wb.close(); raise ValueError("يجب إضافة منتج واحد على الأقل")
                summary = "، ".join(f"{p['product_name']} × {p['quantity']}" for p in products)
                fields["Product_Name"] = summary
                fields["Quantity"] = sum(p["quantity"] for p in products)
                for row in list(wi.iter_rows(min_row=2)):
                    if str(row[1].value) == str(order_id):
                        wi.delete_rows(row[0].row, 1)
                ts = now_str()
                for p in products:
                    wi.append([self._next_item_id(wi), order_id, p["product_name"], p["quantity"], "", "بانتظار التوفر", "", "", "", "", "", "", ts])
            self._update_fields(ws, order_id, fields)
            self._append_log(wl, order_id, "تعديل بيانات الطلب", old, fields.get("Status", old), "تم تعديل بيانات الطلب", user)
            _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu); _atomic_save(wb)
            try:
                _make_backup("auto")
            except Exception:
                pass
        return self.get_order(order_id)

    def _row_snapshot(self, ws, wi, order_id):
        row = self._find_row(ws, order_id)
        if row is None:
            return None
        order = {h: row[i].value if row[i].value is not None else "" for i, h in enumerate(ORDERS_HEADERS)}
        items = []
        for r in wi.iter_rows(min_row=2, values_only=True):
            if str(r[1] or "") == str(order_id):
                items.append({h: (r[i] if i < len(r) and r[i] is not None else "") for i, h in enumerate(ITEM_HEADERS)})
        return {"order": order, "items": items}

    def _invalidate_undo(self, wu, order_id):
        for r in wu.iter_rows(min_row=2):
            if str(r[1].value or "") == str(order_id) and not r[5].value:
                r[5].value = now_str()

    def _add_undo(self, wu, order_id, action, snapshot, user):
        undo_id = f"UNDO-{wu.max_row:06d}"
        wu.append([undo_id, order_id, action, json.dumps(snapshot, ensure_ascii=False), now_str(), "", user or "موظف"])
        return undo_id

    def get_undo_info(self, order_id):
        with _lock:
            wb = self._load()
            if "Undo_History" not in wb.sheetnames:
                wb.close(); return {"available": False}
            wu = wb["Undo_History"]
            rows = []
            for r in wu.iter_rows(min_row=2, values_only=True):
                if not r or str(r[1] or "") != str(order_id):
                    continue
                rows.append({h: (r[i] if i < len(r) and r[i] is not None else "") for i, h in enumerate(UNDO_HEADERS)})
            wb.close()
        active = [x for x in rows if not x.get("Undone_At")]
        if not active:
            return {"available": False}
        active.sort(key=lambda x: str(x.get("Created_At", "")), reverse=True)
        x = active[0]
        return {"available": True, "action": x.get("Action", ""), "created_at": x.get("Created_At", ""), "undo_id": x.get("Undo_ID", "")}

    def undo_last(self, order_id, user="موظف"):
        with _lock:
            wb = self._load()
            if "Undo_History" not in wb.sheetnames:
                wb.close(); return {"error": "ميزة التراجع غير متاحة في ملف البيانات الحالي", "code": 409}
            wu = wb["Undo_History"]
            candidates = []
            for r in wu.iter_rows(min_row=2):
                if str(r[1].value or "") == str(order_id) and not r[5].value:
                    candidates.append(r)
            if not candidates:
                wb.close(); return {"error": "لا يوجد إجراء يمكن التراجع عنه لهذا الطلب", "code": 409}
            target = sorted(candidates, key=lambda r: str(r[4].value or ""), reverse=True)[0]
            snapshot = json.loads(str(target[3].value))
            ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]
            current = self._row_snapshot(ws, wi, order_id)
            if current is None:
                wb.close(); return {"error": "الطلب غير موجود", "code": 404}
            _make_backup("auto")
            order_data = snapshot.get("order", {})
            row = self._find_row(ws, order_id)
            for i, h in enumerate(ORDERS_HEADERS):
                row[i].value = order_data.get(h, "")
            for r in list(wi.iter_rows(min_row=2)):
                if str(r[1].value or "") == str(order_id):
                    wi.delete_rows(r[0].row, 1)
            for item in snapshot.get("items", []):
                wi.append([item.get(h, "") for h in ITEM_HEADERS])
            target[5].value = now_str()
            self._append_log(wl, order_id, f"تراجع عن: {target[2].value}", current["order"].get("Status", ""), order_data.get("Status", ""), "تم التراجع عن آخر تغيير للمستخدم", user)
            _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu)
            _atomic_save(wb)
            try: _make_backup("auto")
            except Exception: pass
        return {"order": self.get_order(order_id), "undone_action": target[2].value}

    def _action(self, order_id, name, allowed_from, fields, note, user):
        with _lock:
            wb = self._load(); ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            old = self._status(ws, order_id)
            if old is None:
                wb.close(); return {"error": "الطلب غير موجود", "code": 404}
            if allowed_from is not None and old not in allowed_from:
                wb.close(); return {"error": f"لا يمكن تنفيذ هذا الإجراء والحالة الحالية هي: {old}", "code": 409}
            snapshot = self._row_snapshot(ws, wi, order_id)
            _make_backup()
            self._invalidate_undo(wu, order_id)
            self._update_fields(ws, order_id, fields)
            new_status = fields.get("Status", old)
            self._append_log(wl, order_id, name, old, new_status, note, user)
            if name in {"تسجيل توفر الطلب", "تم الاتصال بالعميل", "تأجيل المتابعة", "تسليم الطلب للعميل", "إلغاء الطلب"}:
                self._add_undo(wu, order_id, name, snapshot, user)
            _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu); _atomic_save(wb)
            try: _make_backup("auto")
            except Exception: pass
        return {"order": self.get_order(order_id)}

    def set_availability(self, order_id, item_updates, available_date=None, user="موظف"):
        """Set availability/pricing per item, derive order status, and create one undo snapshot."""
        d = available_date or today_str()
        with _lock:
            wb = self._load(); ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            old = self._status(ws, order_id)
            if old is None:
                wb.close(); return {"error": "الطلب غير موجود", "code": 404}
            current_items = self._items_for(self._read_items(wb), order_id)
            if not current_items:
                wb.close(); return {"error": "لا توجد منتجات في هذا الطلب", "code": 409}
            by_id = {str(x.get("Item_ID")): x for x in item_updates if x.get("Item_ID")}
            snapshot = self._row_snapshot(ws, wi, order_id)
            states=[]
            for row in wi.iter_rows(min_row=2):
                if str(row[1].value or "") != str(order_id): continue
                iid=str(row[0].value or "")
                upd=by_id.get(iid, {})
                status=str(upd.get("availability_status") or row[5].value or "بانتظار التوفر").strip()
                if status not in {"متوفر", "غير متوفر", "بانتظار التوفر"}:
                    wb.close(); return {"error": f"حالة توفر غير صحيحة للمنتج {row[2].value}", "code": 400}
                if status == "متوفر":
                    try:
                        normal_raw = str(upd.get("available_price") or "").strip()
                        disc_raw = str(upd.get("discounted_price") or "").strip()
                        normal = float(normal_raw) if normal_raw else None
                        disc = float(disc_raw) if disc_raw else None
                        if normal is not None and normal < 0: raise ValueError
                        if disc is not None and disc < 0: raise ValueError
                        if normal is not None and disc is not None and disc > normal:
                            wb.close(); return {"error": f"سعر الخصم لا يمكن أن يكون أعلى من السعر العادي للمنتج {row[2].value}", "code": 400}
                    except ValueError:
                        wb.close(); return {"error": f"السعر المدخل غير صحيح للمنتج {row[2].value}", "code": 400}
                row[5].value=status
                if status == "متوفر":
                    row[6].value=str(upd.get("available_price") or "").strip()
                    row[7].value=str(upd.get("discounted_price") or "").strip()
                    row[8].value=""
                    row[9].value=str(upd.get("availability_note") or "").strip()
                    row[10].value="نعم" if upd.get("price_confirmation_required") in (True, "true", "True", 1, "1", "نعم") else ""
                    row[11].value=d
                elif status == "غير متوفر":
                    row[6].value=""; row[7].value=""
                    row[8].value=str(upd.get("unavailable_reason") or "").strip()
                    row[9].value=str(upd.get("availability_note") or "").strip()
                    row[10].value=""
                    row[11].value=""
                    if not row[8].value:
                        wb.close(); return {"error": f"يجب اختيار سبب عدم التوفر للمنتج {row[2].value}", "code": 400}
                else:
                    row[6].value=""; row[7].value=""; row[8].value=""; row[9].value=""; row[10].value=""; row[11].value=""
                states.append(status)
            if all(x == "متوفر" for x in states): new_status=STATUS_AVAILABLE
            elif all(x == "غير متوفر" for x in states): new_status=STATUS_UNAVAILABLE
            elif any(x == "متوفر" for x in states) or any(x == "غير متوفر" for x in states): new_status=STATUS_PARTIAL
            else: new_status=STATUS_PENDING
            if new_status == STATUS_PENDING:
                wb.close(); return {"error":"يجب تحديد حالة توفر منتج واحد على الأقل قبل الحفظ", "code":400}
            _make_backup()
            self._invalidate_undo(wu, order_id)
            fields={"Status":new_status, "Available_Date":d if any(x=="متوفر" for x in states) else "", "Next_Followup_Date":today_str() if new_status in (STATUS_AVAILABLE, STATUS_PARTIAL) else ""}
            self._update_fields(ws, order_id, fields)
            note_parts=[]
            for row in wi.iter_rows(min_row=2):
                if str(row[1].value or "") == str(order_id):
                    st=str(row[5].value or "")
                    if st == "متوفر": note_parts.append(f"{row[2].value}: متوفر" + (f" بسعر {row[6].value}" if row[6].value else "") + (f" بعد الخصم {row[7].value}" if row[7].value else ""))
                    elif st == "غير متوفر": note_parts.append(f"{row[2].value}: غير متوفر — {row[8].value}")
            note=" | ".join(note_parts)
            self._append_log(wl, order_id, "تحديث توفر المنتجات", old, new_status, note, user)
            self._add_undo(wu, order_id, "تحديث توفر المنتجات", snapshot, user)
            _format_sheet(ws); _format_sheet(wi); _format_sheet(wl); _format_sheet(wu); _atomic_save(wb)
            try: _make_backup("auto")
            except Exception: pass
        return {"order": self.get_order(order_id)}

    def mark_available(self, order_id, available_date=None, user="موظف"):
        order=self.get_order(order_id)
        if not order: return {"error":"الطلب غير موجود","code":404}
        updates=[{"Item_ID":i.get("Item_ID"),"availability_status":"متوفر"} for i in (order.get("Items") or [])]
        return self.set_availability(order_id, updates, available_date, user)

    def mark_contacted(self, order_id, followup_days=2, user="موظف"):
        contact = today_str(); nxt = add_days(contact, followup_days)
        return self._action(order_id, "تم التواصل مع العميل", {STATUS_AVAILABLE, STATUS_PARTIAL, STATUS_UNAVAILABLE, STATUS_CONTACTED, STATUS_NOT_PICKED}, {"Status": STATUS_CONTACTED, "Contact_Status": CONTACT_AWAITING, "Last_Contact_Date": contact, "Next_Followup_Date": nxt}, f"تم التواصل، بانتظار رد العميل، المتابعة القادمة {nxt}", user)

    def set_contact_status(self, order_id, contact_status, note="", user="موظف"):
        if contact_status not in ALL_CONTACT_STATUSES:
            return {"error": "حالة التواصل غير صحيحة", "code": 400}
        with _lock:
            wb = self._load(); ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            old = self._status(ws, order_id)
            if old is None:
                wb.close(); return {"error": "الطلب غير موجود", "code": 404}
            snapshot = self._row_snapshot(ws, wi, order_id)
            _make_backup(); self._invalidate_undo(wu, order_id)
            fields={"Contact_Status": contact_status}
            if contact_status == CONTACT_ACCEPTED:
                current_items = self._items_for(self._read_items(wb), order_id)
                if not any(str(i.get("Availability_Status") or "") == "متوفر" for i in current_items):
                    wb.close(); return {"error": "لا يمكن تسجيل موافقة العميل لأن لا يوجد منتج متوفر في الطلب", "code": 409}
            if contact_status == CONTACT_AWAITING:
                fields.update({"Last_Contact_Date": today_str(), "Next_Followup_Date": add_days(today_str(), 2)})
            elif contact_status == CONTACT_ACCEPTED:
                fields.update({"Status": STATUS_CONTACTED, "Last_Contact_Date": today_str(), "Next_Followup_Date": ""})
            elif contact_status == CONTACT_REJECTED:
                fields.update({"Status": STATUS_CANCELLED, "Last_Contact_Date": today_str(), "Next_Followup_Date": ""})
            if contact_status == CONTACT_POSTPONED:
                fields.update({"Status": STATUS_NOT_PICKED, "Next_Followup_Date": add_days(today_str(), 1)})
            self._update_fields(ws, order_id, fields)
            self._append_log(wl, order_id, "تحديث حالة التواصل", old, old, note or contact_status, user)
            self._add_undo(wu, order_id, "تحديث حالة التواصل", snapshot, user)
            _format_sheet(ws); _format_sheet(wl); _format_sheet(wu); _atomic_save(wb)
            try: _make_backup("auto")
            except Exception: pass
        return {"order": self.get_order(order_id)}

    def mark_pickup(self, order_id, force=False, user="موظف"):
        order = self.get_order(order_id)
        if not order:
            return {"error": "الطلب غير موجود", "code": 404}
        if order["Status"] in CLOSED_STATUSES and not force:
            return {"error": "هذا الطلب مغلق بالفعل. أكّد العملية للمتابعة.", "code": 409, "needs_confirmation": True}
        return self._action(order_id, "تسليم الطلب للعميل", None, {"Status": STATUS_PICKED_UP, "Pickup_Date": now_str(), "Next_Followup_Date": ""}, "استلم العميل الطلب", user)

    def postpone(self, order_id, days=None, custom_date=None, user="موظف"):
        nxt = custom_date or add_days(today_str(), days or 1)
        return self._action(order_id, "تأجيل المتابعة", {STATUS_CONTACTED, STATUS_NOT_PICKED, STATUS_AVAILABLE}, {"Status": STATUS_NOT_PICKED, "Next_Followup_Date": nxt}, f"تم تأجيل المتابعة إلى {nxt}", user)

    def cancel_order(self, order_id, note="", user="موظف"):
        if not self.get_order(order_id):
            return {"error": "الطلب غير موجود", "code": 404}
        return self._action(order_id, "إلغاء الطلب", None, {"Status": STATUS_CANCELLED, "Next_Followup_Date": ""}, note or "تم إلغاء الطلب", user)

    def delete_order(self, order_id, user="موظف"):
        with _lock:
            _make_backup()
            wb = self._load(); ws = wb["Orders"]; wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            row = self._find_row(ws, order_id)
            if not row:
                wb.close(); return False
            self._invalidate_undo(wu, order_id)
            ws.delete_rows(row[0].row, 1)
            for r in list(wi.iter_rows(min_row=2)):
                if str(r[1].value) == str(order_id):
                    wi.delete_rows(r[0].row, 1)
            self._append_log(wl, order_id, "حذف الطلب نهائيًا", "", "", "تم حذف الطلب من النظام", user)
            _atomic_save(wb)
            self._delete_order_uploads(order_id)
            try:
                _make_backup("auto")
            except Exception:
                pass
        return True

    def _delete_order_uploads(self, order_id):
        folder = os.path.join(UPLOAD_DIR, str(order_id))
        if os.path.isdir(folder):
            shutil.rmtree(folder, ignore_errors=True)

    def set_item_image(self, order_id, item_id, source_stream, filename, content_length=None, user="موظف"):
        ext = os.path.splitext(filename or "")[1].lower()
        if ext not in ALLOWED_IMAGE_EXTS:
            raise ValueError("صيغة الصورة غير مدعومة. استخدم JPG أو PNG أو WEBP")
        if content_length and content_length > MAX_IMAGE_SIZE:
            raise ValueError("حجم الصورة أكبر من 10 ميجابايت")
        data = source_stream.read(MAX_IMAGE_SIZE + 1)
        if len(data) > MAX_IMAGE_SIZE:
            raise ValueError("حجم الصورة أكبر من 10 ميجابايت")
        with _lock:
            wb = self._load(); wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            target_row = None
            for row in wi.iter_rows(min_row=2):
                if str(row[0].value) == str(item_id) and str(row[1].value) == str(order_id):
                    target_row = row
                    break
            if target_row is None:
                wb.close()
                return None
            _make_backup()
            self._invalidate_undo(wu, order_id)
            old_path = str(target_row[4].value or "")
            unique = uuid.uuid4().hex[:10]
            rel_dir = str(order_id)
            abs_dir = os.path.join(UPLOAD_DIR, rel_dir)
            os.makedirs(abs_dir, exist_ok=True)
            rel_path = os.path.join(rel_dir, f"{item_id}_{unique}{ext}").replace(os.sep, "/")
            abs_path = os.path.join(UPLOAD_DIR, rel_path.replace("/", os.sep))
            with open(abs_path, "wb") as f:
                f.write(data)
            target_row[4].value = rel_path
            _format_sheet(wi)
            _format_sheet(wl)
            _format_sheet(wu)
            self._append_log(wl, order_id, "إضافة صورة للمنتج", "", "", f"تم إرفاق صورة بالمنتج {target_row[2].value}", user)
            _atomic_save(wb)
            if old_path and old_path != rel_path:
                try:
                    old_abs = os.path.join(UPLOAD_DIR, old_path.replace("/", os.sep))
                    if os.path.isfile(old_abs): os.remove(old_abs)
                except OSError:
                    pass
            try:
                _make_backup("auto")
            except Exception:
                pass
        return rel_path

    def delete_item_image(self, order_id, item_id, user="موظف"):
        with _lock:
            wb = self._load(); wi = wb["Order_Items"]; wl = wb["Activity_Log"]; wu = wb["Undo_History"]
            target = None
            for row in wi.iter_rows(min_row=2):
                if str(row[0].value) == str(item_id) and str(row[1].value) == str(order_id):
                    target = row; break
            if target is None:
                wb.close(); return False
            old_path = str(target[4].value or "")
            if not old_path:
                wb.close(); return True
            _make_backup()
            self._invalidate_undo(wu, order_id)
            target[4].value = ""
            self._append_log(wl, order_id, "حذف صورة المنتج", "", "", f"تم حذف صورة المنتج {target[2].value}", user)
            _format_sheet(wu)
            _atomic_save(wb)
            try:
                old_abs = os.path.join(UPLOAD_DIR, old_path.replace("/", os.sep))
                if os.path.isfile(old_abs): os.remove(old_abs)
            except OSError:
                pass
            try:
                _make_backup("auto")
            except Exception:
                pass
        return True

    def reset_all_data(self, user="موظف"):
        """Permanently clear all application data for a clean test/reset cycle.
        The program files remain intact. This removes orders, logs, undo history,
        message settings, uploaded images, and backups, then recreates a clean workbook.
        """
        with _lock:
            # Never create a backup here: the purpose is a true clean reset.
            if os.path.isdir(UPLOAD_DIR):
                shutil.rmtree(UPLOAD_DIR, ignore_errors=True)
            if os.path.isdir(BACKUP_DIR):
                shutil.rmtree(BACKUP_DIR, ignore_errors=True)
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            os.makedirs(BACKUP_DIR, exist_ok=True)
            if os.path.isfile(DB_PATH):
                try: os.remove(DB_PATH)
                except OSError as e: return {"error": f"تعذر حذف ملف البيانات: {e}", "code": 500}
            if os.path.isfile(TMP_PATH):
                try: os.remove(TMP_PATH)
                except OSError: pass
            # Reset the optional storage marker as well.
            if os.path.isfile(STORAGE_MARKER):
                try: os.remove(STORAGE_MARKER)
                except OSError: pass
            ensure_db()
            return {"success": True, "message": "تم حذف جميع بيانات التطبيق وإعادته لحالة نظيفة"}

    def storage_info(self):
        return {
            "version": "1.2.0",
            "data_dir": DATA_DIR,
            "backup_dir": BACKUP_DIR,
            "upload_dir": UPLOAD_DIR,
            "shared_storage": True,
            "storage_marker": os.path.exists(STORAGE_MARKER),
        }

    def import_legacy_data(self, source_path):
        """Import a previous Excel/backup ZIP into the current data store safely.
        A backup of the current state is created first. The imported workbook is
        then migrated to the current schema without deleting legacy records.
        """
        source_path = os.path.abspath(source_path)
        if not os.path.isfile(source_path):
            raise ValueError("ملف الاستيراد غير موجود")

        with _lock:
            extract_dir = None
            temp_db = None
            try:
                lower = source_path.lower()
                if lower.endswith('.zip'):
                    extract_dir = tempfile.mkdtemp(prefix='ezz_import_', dir=BASE_DIR)
                    with zipfile.ZipFile(source_path, 'r') as zf:
                        for info in zf.infolist():
                            name = info.filename.replace('\\', '/')
                            if name.startswith('/') or '..' in name.split('/'):
                                raise ValueError("ملف الاستيراد يحتوي مسارًا غير آمن")
                        zf.extractall(extract_dir)
                    candidates = [
                        os.path.join(extract_dir, 'data', 'pharmacy_orders.xlsx'),
                        os.path.join(extract_dir, 'pharmacy_orders.xlsx'),
                    ]
                    src_db = next((x for x in candidates if os.path.isfile(x)), None)
                    if not src_db:
                        raise ValueError("لم يتم العثور على ملف pharmacy_orders.xlsx داخل النسخة")
                elif lower.endswith('.xlsx'):
                    src_db = source_path
                else:
                    raise ValueError("اختر ملف Excel (.xlsx) أو نسخة احتياطية (.zip)")

                # Validate source before touching current data.
                wb = load_workbook(src_db, read_only=True, data_only=False)
                sheets = set(wb.sheetnames)
                if 'Orders' not in sheets:
                    wb.close()
                    raise ValueError("ملف البيانات لا يحتوي ورقة Orders المطلوبة")
                headers = [str(c.value or '').strip() for c in wb['Orders'][1]]
                if not any(h in headers for h in ('Order_ID', 'Customer_Name', 'Phone')):
                    wb.close()
                    raise ValueError("ملف البيانات لا يبدو كملف طلبات صالح")
                wb.close()

                # Preserve current state so import is recoverable.
                backup_name = _make_backup('auto')
                temp_db = DB_PATH + '.import.tmp.xlsx'
                shutil.copy2(src_db, temp_db)
                os.replace(temp_db, DB_PATH)

                # Migrate imported workbook to the current schema.
                ensure_db()

                # Import uploaded product images when a full backup ZIP is used.
                if extract_dir:
                    src_uploads = os.path.join(extract_dir, 'uploads')
                    if os.path.isdir(src_uploads):
                        os.makedirs(UPLOAD_DIR, exist_ok=True)
                        for root, _, files in os.walk(src_uploads):
                            for filename in files:
                                src = os.path.join(root, filename)
                                rel = os.path.relpath(src, src_uploads)
                                dst = os.path.join(UPLOAD_DIR, rel)
                                os.makedirs(os.path.dirname(dst), exist_ok=True)
                                shutil.copy2(src, dst)

                return {"success": True, "backup": backup_name, "order_count": len(self.get_all_orders())}
            except Exception:
                if temp_db and os.path.exists(temp_db):
                    try: os.remove(temp_db)
                    except OSError: pass
                raise
            finally:
                if extract_dir:
                    shutil.rmtree(extract_dir, ignore_errors=True)


    def list_backups(self):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        items=[]
        for f in os.listdir(BACKUP_DIR):
            if not f.lower().endswith((".zip", ".xlsx")): continue
            path=os.path.join(BACKUP_DIR,f); st=os.stat(path)
            reason = "تلقائية" if "_auto_" in f else "يدوية" if "_manual_" in f else "قديمة"
            items.append({"filename":f,"size_kb":round(st.st_size/1024,1),"created_at":datetime.fromtimestamp(st.st_mtime,TZ).strftime("%Y-%m-%d %H:%M:%S"),"reason":reason,"_mtime":st.st_mtime})
        items.sort(key=lambda x:x["_mtime"], reverse=True)
        for x in items: x.pop("_mtime", None)
        return items

    def create_manual_backup(self):
        with _lock: return _make_backup("manual")

    def restore_backup(self, filename):
        safe=os.path.basename(filename); path=os.path.join(BACKUP_DIR,safe)
        if not os.path.exists(path): return False
        with _lock:
            _make_backup("auto")
            if safe.lower().endswith(".zip"):
                extract_dir = os.path.join(DATA_DIR, f"restore_{uuid.uuid4().hex}")
                os.makedirs(extract_dir, exist_ok=True)
                try:
                    with zipfile.ZipFile(path, "r") as zf:
                        for info in zf.infolist():
                            name = info.filename.replace("\\", "/")
                            if name.startswith("/") or ".." in name.split("/"):
                                raise ValueError("ملف النسخة الاحتياطية غير صالح")
                        zf.extractall(extract_dir)
                    db_src = os.path.join(extract_dir, "data", "pharmacy_orders.xlsx")
                    uploads_src = os.path.join(extract_dir, "uploads")
                    if not os.path.isfile(db_src):
                        raise ValueError("النسخة الاحتياطية لا تحتوي على ملف البيانات")
                    shutil.copy2(db_src, TMP_PATH)
                    os.replace(TMP_PATH, DB_PATH)
                    if os.path.isdir(UPLOAD_DIR): shutil.rmtree(UPLOAD_DIR, ignore_errors=True)
                    if os.path.isdir(uploads_src): shutil.copytree(uploads_src, UPLOAD_DIR)
                    else: os.makedirs(UPLOAD_DIR, exist_ok=True)
                finally:
                    shutil.rmtree(extract_dir, ignore_errors=True)
            else:
                shutil.copy2(path, TMP_PATH); os.replace(TMP_PATH, DB_PATH)
                os.makedirs(UPLOAD_DIR, exist_ok=True)
            ensure_db()
            try:
                _make_backup("auto")
            except Exception:
                pass
        return True


def _sheet_headers_compatible(ws, wanted):
    current = [str(c.value or "") for c in ws[1]]
    # For old sheets, use the known ordering and fill missing fields from blanks.
    if current == wanted:
        return wanted
    result = []
    for h in wanted:
        result.append(h)
    return result


# Backend selection:
# - Local Windows/Linux use the durable Excel backend by default.
# - Cloud deployment switches to PostgreSQL/Supabase when DATABASE_URL is set.
if os.environ.get("DATABASE_URL", "").strip():
    from cloud_db import CloudDB
    db = CloudDB()
else:
    db = ExcelDB()
