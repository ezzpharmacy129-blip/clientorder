# -*- coding: utf-8 -*-
"""Runtime patch loaded automatically by Python before Gunicorn imports app."""

from io import BytesIO
from datetime import datetime


def _install_export_route():
    from app import app, db, list_pharmacy_shortages
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    endpoint = "/api/data/export-postrollback"

    # Avoid duplicate registration if this module is reloaded.
    if any(rule.rule == endpoint for rule in app.url_map.iter_rules()):
        return

    @app.get(endpoint)
    def api_export_postrollback():
        """Download current application data as an Excel workbook."""
        wb = Workbook()

        ws = wb.active
        ws.title = "Orders"
        orders = db.get_all_orders()
        order_headers = [
            "Order_ID", "Customer_Name", "Phone", "Product_Name", "Quantity",
            "Order_Date", "Available_Date", "Status", "Contact_Status",
            "Last_Contact_Date", "Next_Followup_Date", "Pickup_Date", "Notes",
            "Created_At", "Updated_At",
        ]
        ws.append(order_headers)
        for order in orders:
            ws.append([order.get(h, "") for h in order_headers])

        ws_items = wb.create_sheet("Order_Items")
        item_headers = [
            "Item_ID", "Order_ID", "Product_Name", "Quantity", "Image_Path",
            "Availability_Status", "Available_Price", "Discounted_Price",
            "Unavailable_Reason", "Availability_Note",
            "Price_Confirmation_Required", "Available_At", "Created_At",
        ]
        ws_items.append(item_headers)
        for order in orders:
            for item in (order.get("Items") or []):
                ws_items.append([item.get(h, "") for h in item_headers])

        ws_short = wb.create_sheet("Pharmacy_Shortages")
        try:
            shortages = list_pharmacy_shortages()
        except Exception:
            shortages = []
        short_headers = [
            "shortage_id", "product_name", "quantity", "note", "status",
            "created_at", "updated_at", "created_by", "resolved_at",
        ]
        ws_short.append(short_headers)
        for row in shortages:
            ws_short.append([row.get(h, "") for h in short_headers])

        ws_users = wb.create_sheet("Users")
        user_headers = ["user_id", "username", "name", "role", "active", "created_at", "last_login"]
        ws_users.append(user_headers)
        try:
            users = db.get_all_users()
        except Exception:
            users = []
        for user in users:
            ws_users.append([user.get(h, "") for h in user_headers])

        ws_log = wb.create_sheet("Activity_Log")
        log_headers = ["log_id", "order_id", "action", "old_status", "new_status", "note", "created_at", "user_name"]
        ws_log.append(log_headers)
        try:
            rows = db.get_all_activity_log()
        except Exception:
            rows = []
        for row in rows:
            ws_log.append([row.get(h, "") for h in log_headers])

        ws_undo = wb.create_sheet("Undo_History")
        undo_headers = ["undo_id", "order_id", "action", "snapshot_json", "created_at", "undone_at", "user_name"]
        ws_undo.append(undo_headers)
        try:
            rows = db.get_all_undo_history()
        except Exception:
            rows = []
        for row in rows:
            ws_undo.append([row.get(h, "") for h in undo_headers])

        ws_set = wb.create_sheet("Settings")
        ws_set.append(["key", "value"])
        try:
            settings = db.get_settings()
        except Exception:
            settings = {}
        for key, value in settings.items():
            ws_set.append([key, value])

        for sheet in wb.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for col_cells in sheet.columns:
                values = [str(cell.value or "") for cell in col_cells]
                width = min(max(len(v) for v in values) + 2, 50)
                sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = max(width, 10)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        return send_file(
            output,
            as_attachment=True,
            download_name=f"ezz-pharmacy-export-{stamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


try:
    _install_export_route()
except Exception:
    # Keep normal application startup unaffected if the optional patch fails.
    pass
