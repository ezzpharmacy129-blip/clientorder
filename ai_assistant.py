# -*- coding: utf-8 -*-
"""Protected OpenAI document/image assistant for Ezz Pharmacy."""
import base64
import io
import json
import os
import threading
import time
from collections import defaultdict, deque

from flask import jsonify, request, render_template_string

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

MAX_AI_FILE_SIZE = 10 * 1024 * 1024
MAX_AI_TEXT_CHARS = 120_000
IMAGE_EXT = {"png", "jpg", "jpeg", "webp", "gif"}
TEXT_EXT = {"txt", "md", "csv", "json", "xlsx"}
FILE_EXT = {"pdf", "docx"}

_rate_lock = threading.Lock()
_rate = defaultdict(deque)
_RATE_WINDOW = 3600
_RATE_MAX = 20


def _ext(name):
    name = str(name or "").lower().rsplit("/", 1)[-1]
    return name.rsplit(".", 1)[-1] if "." in name else ""


def _rate_ok(key):
    now = time.time()
    with _rate_lock:
        q = _rate[key]
        while q and now - q[0] > _RATE_WINDOW:
            q.popleft()
        if len(q) >= _RATE_MAX:
            return False
        q.append(now)
        return True


def _xlsx_text(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    try:
        for ws in wb.worksheets:
            out.append("ورقة: " + ws.title)
            for n, row in enumerate(ws.iter_rows(values_only=True), 1):
                vals = ["" if v is None else str(v) for v in row[:30]]
                if any(vals):
                    out.append(" | ".join(vals))
                if n >= 500:
                    out.append("… تم اختصار الورقة بعد 500 صف.")
                    break
    finally:
        wb.close()
    return "\n".join(out)


def _text_content(ext, data):
    if ext == "xlsx":
        return _xlsx_text(data)
    if ext in {"txt", "md", "csv"}:
        return data.decode("utf-8-sig", errors="replace")
    if ext == "json":
        raw = data.decode("utf-8-sig", errors="replace")
        try:
            return json.dumps(json.loads(raw), ensure_ascii=False, indent=2)
        except Exception:
            return raw
    return ""


def install_ai(app):
    if getattr(app, "_ezz_ai_installed", False):
        return
    app._ezz_ai_installed = True

    def current_user():
        getter = app.extensions.get("ezz_auth", {}).get("current_user")
        if getter:
            try:
                u = getter()
                if u:
                    return u
            except Exception:
                pass
        return None

    @app.get("/ai")
    def ai_page():
        if not current_user():
            return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        return render_template_string(r"""<!doctype html>
<html lang="ar" dir="rtl"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>مساعد الذكاء الاصطناعي - صيدلية عز الصحة</title>
<style>
:root{--p:#20a6b1;--pd:#0c6d78;--n:#153f7a;--bg:#f4f8fb;--b:#dce8ee;--t:#17324d;--m:#71818d}
*{box-sizing:border-box}body{margin:0;background:linear-gradient(180deg,#f8fbfd,#f4f8fb);font-family:Tahoma,Arial,sans-serif;color:var(--t)}
.wrap{max-width:980px;margin:auto;padding:24px}.top{display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap;margin-bottom:18px}
.brand{display:flex;gap:12px;align-items:center}.brand img{width:58px;height:58px;object-fit:contain;background:#fff;border:1px solid var(--b);border-radius:15px;padding:5px}
h1{margin:0;color:var(--n);font-size:22px}.sub{margin:4px 0 0;color:var(--m);font-size:12px}
.card{background:#fff;border:1px solid var(--b);border-radius:18px;padding:20px;box-shadow:0 10px 30px rgba(21,63,122,.08);margin-bottom:16px}
label{display:block;font-weight:800;font-size:13px;margin-bottom:7px}
input,textarea{width:100%;font:inherit;border:1px solid var(--b);border-radius:11px;padding:11px;background:#fff;color:var(--t)}
textarea{min-height:120px;resize:vertical}.filebox{padding:15px;border:1px dashed #8dcdd3;border-radius:14px;background:#f7fcfd}
.hint,.status{font-size:12px;color:var(--m);line-height:1.7;margin-top:7px}.actions{display:flex;gap:9px;flex-wrap:wrap;margin-top:14px}
.btn{border:0;border-radius:11px;padding:10px 16px;font-weight:800;cursor:pointer;text-decoration:none}.primary{background:linear-gradient(135deg,var(--p),#148f9b);color:#fff}.secondary{background:#eef4f6;color:var(--t)}
.btn:disabled{opacity:.55;cursor:not-allowed}.result{white-space:pre-wrap;line-height:1.9;font-size:14px;background:#fbfeff;border:1px solid var(--b);border-radius:12px;padding:15px;min-height:160px}
.notice{background:#fff8df;border:1px solid #f1df94;color:#6d5a18;padding:12px 14px;border-radius:11px;font-size:12px;line-height:1.8}.error{color:#a12d2d}
.back{color:var(--pd);font-weight:800;text-decoration:none}
</style></head>
<body><div class="wrap">
<div class="top"><div class="brand"><img src="/static/logo-mark.png" alt="صيدلية عز الصحة"><div><h1>🤖 مساعد الذكاء الاصطناعي</h1><p class="sub">قراءة وتحليل الصور والمستندات داخل نظام عز الصحة</p></div></div><a class="back" href="/">← العودة للنظام</a></div>
<div class="card">
<label for="ai-file">الملف أو الصورة</label><div class="filebox"><input id="ai-file" type="file" accept=".png,.jpg,.jpeg,.webp,.gif,.pdf,.txt,.md,.csv,.json,.docx,.xlsx"><div class="hint">صورة، PDF، Word، Excel، CSV أو ملف نصي — الحد الأقصى 10 ميجابايت.</div></div>
<div style="margin-top:16px"><label for="ai-prompt">ماذا تريد مني أن أقرأ أو أستخرج؟</label><textarea id="ai-prompt" placeholder="مثال: استخرج أسماء المنتجات والكميات والأسعار ورتبها في جدول واضح."></textarea></div>
<div class="notice" style="margin-top:14px">المساعد لتحليل المستندات والبيانات ومساعدة الموظف. يجب مراجعة أي نتيجة متعلقة بالدواء أو المريض من الموظف المختص قبل اعتمادها.</div>
<div class="actions"><button id="ai-submit" class="btn primary">🤖 تحليل الآن</button><button id="ai-clear" class="btn secondary" type="button">مسح</button></div>
<div id="ai-status" class="status"></div></div>
<div class="card"><div style="display:flex;justify-content:space-between;align-items:center;gap:10px;margin-bottom:10px"><strong style="color:var(--n)">النتيجة</strong><button id="ai-copy" class="btn secondary" type="button">📋 نسخ النتيجة</button></div><div id="ai-result" class="result">لم يتم تحليل أي ملف بعد.</div></div>
</div>
<script>
const f=document.getElementById('ai-file'),p=document.getElementById('ai-prompt'),s=document.getElementById('ai-submit'),c=document.getElementById('ai-clear'),cp=document.getElementById('ai-copy'),st=document.getElementById('ai-status'),r=document.getElementById('ai-result');
s.onclick=async()=>{const file=f.files&&f.files[0],prompt=p.value.trim();if(!file&&!prompt){st.textContent='اختر ملفًا أو اكتب المطلوب.';st.className='status error';return}
s.disabled=true;st.textContent='جاري القراءة والتحليل...';st.className='status';r.textContent='جارٍ التحليل...';const fd=new FormData();if(file)fd.append('file',file);fd.append('prompt',prompt);
try{const x=await fetch('/api/ai/analyze',{method:'POST',body:fd,credentials:'same-origin'}),d=await x.json();if(!x.ok)throw new Error(d.error||'تعذر تنفيذ التحليل');r.textContent=d.result||'لم تُرجع نتيجة.';st.textContent=d.filename?'تم تحليل: '+d.filename:'تم التحليل بنجاح ✅';}catch(e){r.textContent='';st.textContent=e.message;st.className='status error'}finally{s.disabled=false}};
c.onclick=()=>{f.value='';p.value='';r.textContent='لم يتم تحليل أي ملف بعد.';st.textContent='';};
cp.onclick=()=>navigator.clipboard?.writeText(r.textContent||'').then(()=>{st.textContent='تم نسخ النتيجة ✅'});
</script></body></html>""")

    @app.post("/api/ai/analyze")
    def ai_analyze():
        u = current_user()
        if not u:
            return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        if OpenAI is None:
            return jsonify({"error": "مكتبة OpenAI غير مثبتة على الخادم"}), 503
        key = str(u.get("username") or u.get("user_id") or "user")
        if not _rate_ok(key):
            return jsonify({"error": "تم الوصول للحد المؤقت لاستخدام المساعد. حاول لاحقًا."}), 429
        api_key = os.environ.get("OPENAI_API_KEY", "").strip()
        if not api_key:
            return jsonify({"error": "لم يتم إعداد OPENAI_API_KEY في Render بعد"}), 503

        prompt = str(request.form.get("prompt") or "").strip()
        if not prompt:
            prompt = "اقرأ الملف المرفق واستخرج أهم المعلومات بشكل منظم وواضح بالعربية. لا تخمّن المعلومات غير الموجودة."

        upload = request.files.get("file")
        filename = str(upload.filename or "") if upload else ""
        data = upload.read(MAX_AI_FILE_SIZE + 1) if upload else b""
        if len(data) > MAX_AI_FILE_SIZE:
            return jsonify({"error": "حجم الملف أكبر من 10 ميجابايت"}), 413

        ext = _ext(filename)
        if filename and ext not in IMAGE_EXT | TEXT_EXT | FILE_EXT:
            return jsonify({"error": "نوع الملف غير مدعوم"}), 400

        client = OpenAI(api_key=api_key)
        model = os.environ.get("OPENAI_MODEL", "gpt-5.6-luna")
        try:
            if not filename:
                response = client.responses.create(model=model, input=prompt)
            elif ext in IMAGE_EXT:
                mime = {"png":"image/png","jpg":"image/jpeg","jpeg":"image/jpeg","webp":"image/webp","gif":"image/gif"}[ext]
                encoded = base64.b64encode(data).decode("ascii")
                response = client.responses.create(model=model, input=[{"role":"user","content":[
                    {"type":"input_text","text":prompt},
                    {"type":"input_image","image_url":"data:"+mime+";base64,"+encoded}
                ]}])
            elif ext in TEXT_EXT:
                content = _text_content(ext, data)[:MAX_AI_TEXT_CHARS]
                response = client.responses.create(model=model, input=[{"role":"user","content":[
                    {"type":"input_text","text":prompt},
                    {"type":"input_text","text":"محتوى الملف ("+filename+"):\n"+content}
                ]}])
            else:
                mime = "application/pdf" if ext == "pdf" else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                encoded = base64.b64encode(data).decode("ascii")
                response = client.responses.create(model=model, input=[{"role":"user","content":[
                    {"type":"input_text","text":prompt},
                    {"type":"input_file","filename":filename,"file_data":"data:"+mime+";base64,"+encoded}
                ]}])

            result = getattr(response, "output_text", "") or "تمت المعالجة لكن لم تُرجع الخدمة نصًا قابلًا للعرض."
            try:
                audit = app.extensions.get("ezz_auth", {}).get("audit")
                if audit:
                    audit(action="AI Analysis", note="تحليل ملف عبر المساعد: "+(filename or "نص مباشر"))
            except Exception:
                pass
            return jsonify({"success": True, "result": result, "filename": filename or None})
        except Exception as exc:
            app.logger.exception("OpenAI analysis failed")
            return jsonify({"error": "تعذر تنفيذ تحليل الذكاء الاصطناعي"}), 500


