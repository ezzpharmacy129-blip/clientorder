# -*- coding: utf-8 -*-
"""Database-backed authentication layer for the existing Excel-backed app."""
import os, sqlite3, secrets, hashlib, hmac, base64, uuid, html
from datetime import datetime, timedelta
from functools import wraps
from zoneinfo import ZoneInfo
from flask import request, session, redirect, jsonify, url_for, render_template_string

TZ = ZoneInfo("Asia/Riyadh")

def now_str(): return datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

def hash_password(password):
    password = str(password or "")
    if not password: raise ValueError("كلمة المرور مطلوبة")
    iterations = 310000; salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, iterations)
    enc=lambda b: base64.urlsafe_b64encode(b).decode().rstrip("=")
    return f"pbkdf2_sha256${iterations}${enc(salt)}${enc(digest)}"

def verify_password(password, encoded):
    try:
        method, it, salt, digest = str(encoded).split("$",3)
        if method != "pbkdf2_sha256": return False
        dec=lambda s: base64.urlsafe_b64decode(s+"="*(-len(s)%4))
        actual=hashlib.pbkdf2_hmac("sha256",str(password).encode(),dec(salt),int(it))
        return hmac.compare_digest(actual,dec(digest))
    except Exception: return False

def install_auth(app, db):
    if getattr(app,"_ezz_auth_installed",False): return
    app._ezz_auth_installed=True
    secret=os.environ.get("SECRET_KEY","").strip()
    if not secret:
        if os.environ.get("RENDER") or os.environ.get("PRODUCTION"): raise RuntimeError("SECRET_KEY is required in production")
        secret=secrets.token_hex(32)
    app.secret_key=secret
    app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE="Lax",SESSION_COOKIE_SECURE=bool(os.environ.get("RENDER") or os.environ.get("PRODUCTION")),PERMANENT_SESSION_LIFETIME=timedelta(hours=12))

    root=os.path.abspath(getattr(db,"SHARED_ROOT",os.path.join(os.path.expanduser("~"),".ezz_pharmacy_fresh")))
    os.makedirs(root,exist_ok=True); users_db=os.path.join(root,"users.sqlite3")
    def conn():
        c=sqlite3.connect(users_db,timeout=20); c.row_factory=sqlite3.Row; return c
    def ensure_users():
        with conn() as c:
            c.execute("""CREATE TABLE IF NOT EXISTS users(
                user_id TEXT PRIMARY KEY, username TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
                password_hash TEXT NOT NULL, role TEXT NOT NULL CHECK(role IN ('admin','employee')),
                active INTEGER NOT NULL DEFAULT 1, created_at TEXT NOT NULL)""")
            au=os.environ.get("ADMIN_USERNAME","").strip(); ap=os.environ.get("ADMIN_PASSWORD","")
            if au and ap and c.execute("SELECT 1 FROM users WHERE username=?",(au,)).fetchone() is None:
                c.execute("INSERT INTO users VALUES(?,?,?,?,?,?,?)",(str(uuid.uuid4()),au,au,hash_password(ap),"admin",1,now_str()))
    def get_user(user_id=None,username=None):
        with conn() as c:
            row=c.execute("SELECT user_id,username,name,password_hash,role,active,created_at FROM users WHERE user_id=?" if user_id else "SELECT user_id,username,name,password_hash,role,active,created_at FROM users WHERE username=?",(user_id or username,)).fetchone()
        return dict(row) if row else None
    def current_user():
        u=get_user(user_id=session.get("user_id")) if session.get("user_id") else None
        if not u or not u["active"]:
            if session.get("user_id"): session.clear()
            return None
        return u
    def audit(order_id="",action="",old_status="",new_status="",note="",actor=None):
        actor=actor or current_user(); username=actor["name"] if actor else "غير مسجل"
        try:
            from openpyxl import load_workbook
            path=getattr(db,"DB_PATH","")
            if not path or not os.path.exists(path): return
            lock=getattr(db,"_lock",None)
            if lock: lock.acquire()
            try:
                wb=load_workbook(path); ws=wb["Activity_Log"] if "Activity_Log" in wb.sheetnames else wb.create_sheet("Activity_Log")
                if ws.max_row==1 and not ws.cell(1,1).value: ws.append(["Log_ID","Order_ID","Action","Old_Status","New_Status","Note","Created_At","User"])
                ws.append(["AUTH-"+uuid.uuid4().hex[:12],str(order_id or ""),action,old_status or "",new_status or "",note or "",now_str(),username]); wb.save(path); wb.close()
            finally:
                if lock: lock.release()
        except Exception: pass
    ensure_users()
    app.extensions["ezz_auth"]={"current_user":current_user,"get_user":get_user,"audit":audit,"hash_password":hash_password,"verify_password":verify_password}
    db._auth_user_provider=current_user

    # Replace the existing default "موظف" actor with the authenticated user's real name.
    for attr in ("_log","_append_log"):
        original=getattr(db,attr,None)
        if original and not getattr(original,"_ezz_wrapped",False):
            def make_wrapper(fn):
                @wraps(fn)
                def wrapped(*args,**kwargs):
                    u=current_user()
                    if u and kwargs.get("user","موظف")=="موظف": kwargs["user"]=u["name"]
                    return fn(*args,**kwargs)
                wrapped._ezz_wrapped=True; return wrapped
            setattr(db,attr,make_wrapper(original))

    @app.before_request
    def require_auth():
        p=request.path
        if p in ("/login","/logout","/health") or p.startswith("/static/"): return None
        if current_user(): return None
        if p.startswith("/api/") or p.startswith("/uploads/"): return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}),401
        return redirect(url_for("ezz_login",next=request.full_path))

    @app.route("/login",methods=["GET","POST"],endpoint="ezz_login")
    def login():
        if request.method=="GET": return redirect(url_for("index")) if current_user() else login_page()
        username=str(request.form.get("username") or "").strip(); password=str(request.form.get("password") or ""); u=get_user(username=username)
        if u and u["active"] and verify_password(password,u["password_hash"]):
            session.clear(); session.permanent=True; session["user_id"]=u["user_id"]; session["username"]=u["username"]; session["role"]=u["role"]; audit(action="Login",note="تسجيل دخول ناجح",actor=u)
            nxt=request.args.get("next") or url_for("index"); return redirect(nxt if str(nxt).startswith("/") else url_for("index"))
        audit(action="Failed Login",note=f"محاولة دخول فاشلة باسم المستخدم: {username or 'غير معروف'}",actor={"name":username or "غير معروف"}); return login_page("اسم المستخدم أو كلمة المرور غير صحيحة."),401

    @app.route("/logout",methods=["GET","POST"],endpoint="ezz_logout")
    def logout():
        u=current_user()
        if u: audit(action="Logout",note="تسجيل الخروج",actor=u)
        session.clear(); r=redirect(url_for("ezz_login")); r.headers["Cache-Control"]="no-store"; return r

    def admin_only(fn):
        @wraps(fn)
        def wrapped(*a,**kw):
            u=current_user()
            if not u:return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}),401
            if u["role"]!="admin":return jsonify({"error":"غير مصرح لك بهذا الإجراء"}),403
            return fn(*a,**kw)
        return wrapped
    @app.get("/api/auth/me")
    def auth_me():
        u=current_user(); return jsonify({"authenticated":bool(u),"user":({"user_id":u["user_id"],"username":u["username"],"name":u["name"],"role":u["role"]} if u else None)})
    @app.get("/admin")
    @admin_only
    def admin_dashboard():
        with conn() as c:n=c.execute("SELECT COUNT(*) c FROM users WHERE active=1").fetchone()["c"]
        return admin_html("لوحة الإدارة",f"<div class='admin-card'><strong>{n}</strong><span>المستخدمون النشطون</span></div><div class='admin-actions'><a href='/admin/users'>إدارة المستخدمين</a><a href='/admin/audit'>Audit Log</a><a href='/'>العودة للنظام</a></div>")
    @app.get("/admin/users")
    @admin_only
    def admin_users():
        with conn() as c: rows=c.execute("SELECT user_id,username,name,role,active FROM users ORDER BY created_at DESC").fetchall()
        body="<div class='admin-actions'><a href='/admin'>لوحة الإدارة</a><a href='/admin/audit'>Audit Log</a><a href='/'>العودة للنظام</a></div><div class='panel'><h2>إضافة مستخدم</h2><form id='add-user' class='form-grid'><input name='name' placeholder='الاسم' required><input name='username' placeholder='Username' required><input name='password' type='password' placeholder='كلمة المرور' required><select name='role'><option value='employee'>employee</option><option value='admin'>admin</option></select><button>إضافة مستخدم</button></form></div><div class='panel'><h2>المستخدمون</h2><table><tr><th>الاسم</th><th>Username</th><th>الدور</th><th>الحالة</th><th>الإجراءات</th></tr>"
        for r in rows: body+=f"<tr><td>{esc(r['name'])}</td><td>{esc(r['username'])}</td><td>{esc(r['role'])}</td><td>{'نشط' if r['active'] else 'معطل'}</td><td><button onclick=\"toggleUser('{r['user_id']}')\">{'تعطيل' if r['active'] else 'تفعيل'}</button> <button onclick=\"changePassword('{r['user_id']}')\">تغيير كلمة المرور</button></td></tr>"
        body+="</table></div><script>async function j(u,b){let r=await fetch(u,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b||{})});let d=await r.json();if(!r.ok)throw Error(d.error||'حدث خطأ');return d}document.getElementById('add-user').onsubmit=async e=>{e.preventDefault();let f=e.target;try{await j('/api/admin/users',{name:f.name.value,username:f.username.value,password:f.password.value,role:f.role.value});location.reload()}catch(x){alert(x.message)}};async function toggleUser(id){try{await j('/api/admin/users/'+id+'/toggle')}catch(x){alert(x.message)}location.reload()}async function changePassword(id){let p=prompt('كلمة المرور الجديدة');if(!p)return;try{await j('/api/admin/users/'+id+'/password',{password:p});alert('تم تغيير كلمة المرور')}catch(x){alert(x.message)}}</script>"
        return admin_html("إدارة المستخدمين",body)
    @app.get("/admin/audit")
    @admin_only
    def admin_audit():
        rows=[]
        try:
            from openpyxl import load_workbook; path=getattr(db,"DB_PATH","")
            if os.path.exists(path):
                wb=load_workbook(path,read_only=True,data_only=True); ws=wb["Activity_Log"] if "Activity_Log" in wb.sheetnames else None
                if ws: rows=list(ws.iter_rows(min_row=2,values_only=True))[-1000:][::-1]
                wb.close()
        except Exception: pass
        body="<div class='admin-actions'><a href='/admin'>لوحة الإدارة</a><a href='/admin/users'>إدارة المستخدمين</a><a href='/'>العودة للنظام</a></div><div class='panel'><h2>Audit Log</h2><table><tr><th>التاريخ</th><th>المستخدم</th><th>العملية</th><th>الطلب</th><th>التفاصيل</th></tr>"
        for r in rows:
            r=list(r)+[""]*8; detail=str(r[5] or "")
            if r[3] or r[4]: detail+=(" — " if detail else "")+f"{r[3] or '—'} ← {r[4] or '—'}"
            body+=f"<tr><td>{esc(r[6])}</td><td>{esc(r[7])}</td><td>{esc(r[2])}</td><td>{esc(r[1])}</td><td>{esc(detail)}</td></tr>"
        return admin_html("Audit Log",body+"</table></div>")
    @app.post("/api/admin/users")
    @admin_only
    def api_add_user():
        d=request.get_json(silent=True) or {}; name=str(d.get("name") or "").strip(); username=str(d.get("username") or "").strip(); password=str(d.get("password") or ""); role=str(d.get("role") or "employee")
        if not name or len(username)<3 or len(password)<8:return jsonify({"error":"الاسم مطلوب واسم المستخدم 3 أحرف على الأقل وكلمة المرور 8 أحرف على الأقل"}),400
        if role not in ("admin","employee"):return jsonify({"error":"الدور غير صحيح"}),400
        with conn() as c:
            try:c.execute("INSERT INTO users VALUES(?,?,?,?,?,?,?)",(str(uuid.uuid4()),username,name,hash_password(password),role,1,now_str()))
            except sqlite3.IntegrityError:return jsonify({"error":"اسم المستخدم موجود بالفعل"}),409
        audit(action="Create User",note=f"إنشاء المستخدم {username}");return jsonify({"success":True}),201
    @app.post("/api/admin/users/<uid>/toggle")
    @admin_only
    def api_toggle_user(uid):
        actor=current_user()
        with conn() as c:
            r=c.execute("SELECT username,active FROM users WHERE user_id=?",(uid,)).fetchone()
            if not r:return jsonify({"error":"المستخدم غير موجود"}),404
            if uid==actor["user_id"] and r["active"]:return jsonify({"error":"لا يمكنك تعطيل حسابك الحالي"}),400
            new=0 if r["active"] else 1;c.execute("UPDATE users SET active=? WHERE user_id=?",(new,uid))
        audit(action="Toggle User",note=f"تغيير حالة المستخدم {r['username']} إلى {'نشط' if new else 'معطل'}");return jsonify({"success":True,"active":bool(new)})
    @app.post("/api/admin/users/<uid>/password")
    @admin_only
    def api_change_password(uid):
        p=str((request.get_json(silent=True) or {}).get("password") or "")
        if len(p)<8:return jsonify({"error":"كلمة المرور يجب ألا تقل عن 8 أحرف"}),400
        with conn() as c:
            r=c.execute("SELECT username FROM users WHERE user_id=?",(uid,)).fetchone()
            if not r:return jsonify({"error":"المستخدم غير موجود"}),404
            c.execute("UPDATE users SET password_hash=? WHERE user_id=?",(hash_password(p),uid))
        audit(action="Change Password",note=f"تغيير كلمة مرور المستخدم {r['username']}");return jsonify({"success":True})
    @app.before_request
    def admin_destructive_guard():
        if request.path in ("/api/data/reset","/api/backups/restore"):
            u=current_user()
            if not u:return jsonify({"error":"تسجيل الدخول مطلوب","authenticated":False}),401
            if u["role"]!="admin":return jsonify({"error":"غير مصرح لك بهذا الإجراء"}),403
    @app.after_request
    def auth_headers(response):
        response.headers.setdefault("Cache-Control","no-store"); response.headers.setdefault("X-Content-Type-Options","nosniff"); response.headers.setdefault("X-Frame-Options","SAMEORIGIN"); response.headers.setdefault("Referrer-Policy","same-origin")
        if response.mimetype=="text/html" and request.path!="/login" and current_user():
            try:
                text=response.get_data(as_text=True); marker="</header>"
                if marker in text and "auth-user-bar" not in text:
                    u=current_user(); admin="<a href='/admin'>الإدارة</a>" if u["role"]=="admin" else ""; bar=f"<div class='auth-user-bar'><span>المستخدم: <strong>{esc(u['name'])}</strong></span>{admin}<a href='/logout'>تسجيل الخروج</a></div>"; response.set_data(text.replace(marker,bar+marker,1))
            except Exception: pass
        return response

def esc(v): return html.escape(str(v or ""),quote=True)
def login_page(error=""):
    e=f"<div class='error'>{esc(error)}</div>" if error else ""
    return render_template_string("""<!doctype html><html lang='ar' dir='rtl'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>تسجيل الدخول - صيدلية عز الصحة</title><style>body{margin:0;min-height:100vh;display:grid;place-items:center;background:#f5fafb;font-family:Tahoma,Arial,sans-serif;color:#17324d}.card{width:min(420px,92vw);background:#fff;border:1px solid #d9e7ea;border-radius:22px;padding:30px;box-shadow:0 18px 60px rgba(18,63,122,.14)}h1{text-align:center;color:#123f7a;font-size:23px}.field{margin-bottom:14px}.field label{display:block;font-weight:700;margin-bottom:7px}.field input{width:100%;padding:13px;box-sizing:border-box;border:1px solid #d9e7ea;border-radius:11px;font-size:16px}.submit{width:100%;border:0;border-radius:11px;padding:13px;background:#0b8f9b;color:#fff;font-weight:800;font-size:16px}.error{background:#fff0f0;color:#a62b2b;border:1px solid #f0c8c8;padding:10px;border-radius:10px;margin-bottom:12px}</style></head><body><div class='card'><h1>صيدلية عز الصحة</h1><p style='text-align:center'>تسجيل الدخول إلى نظام متابعة الطلبات</p>__ERROR__<form method='post'><div class='field'><label>اسم المستخدم</label><input name='username' autocomplete='username' required autofocus></div><div class='field'><label>كلمة المرور</label><input type='password' name='password' autocomplete='current-password' required></div><button class='submit'>تسجيل الدخول</button></form></div></body></html>""".replace("__ERROR__",e))
def admin_html(title,body):
    return render_template_string("""<!doctype html><html lang='ar' dir='rtl'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>{{title}} - صيدلية عز الصحة</title><style>:root{--primary:#0b8f9b;--navy:#123f7a;--border:#d9e7ea;--bg:#f5fafb}*{box-sizing:border-box}body{margin:0;background:var(--bg);color:#17324d;font-family:Tahoma,Arial,sans-serif}.wrap{max-width:1200px;margin:auto;padding:22px}.top{display:flex;justify-content:space-between;align-items:center;margin-bottom:18px}.top a,.admin-actions a{display:inline-block;text-decoration:none;color:var(--navy);background:#fff;border:1px solid var(--border);padding:9px 12px;border-radius:9px;font-weight:700}.admin-card,.panel{background:#fff;border:1px solid var(--border);border-radius:14px;padding:18px;margin-bottom:16px}.admin-card strong{display:block;font-size:30px;color:var(--navy)}.admin-card span{color:#6c7f8a;font-size:13px}.admin-actions{display:flex;gap:8px;flex-wrap:wrap;margin:12px 0 18px}.form-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:8px}.form-grid input,.form-grid select{padding:10px;border:1px solid var(--border);border-radius:9px}.form-grid button{border:0;border-radius:9px;background:var(--primary);color:#fff;font-weight:700;padding:10px}table{width:100%;border-collapse:collapse;min-width:700px}th,td{padding:10px;border-bottom:1px solid var(--border);text-align:right;font-size:13px}th{background:#f4fafb;color:var(--navy)}@media(max-width:760px){.form-grid{grid-template-columns:1fr}.wrap{padding:14px}}</style></head><body><div class='wrap'><div class='top'><h1>{{title}}</h1><a href='/logout'>تسجيل الخروج</a></div>{{body|safe}}</div></body></html>""",title=title,body=body)
