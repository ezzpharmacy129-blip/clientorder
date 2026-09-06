# -*- coding: utf-8 -*-
"""PostgreSQL/Supabase data layer for the cloud deployment."""
import io
import json
import os
import re
import uuid
import zipfile
from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
import psycopg
from psycopg.rows import dict_row

from workflow_policy import validate_pickup, validate_not_picked
from db import (
    ALL_CONTACT_STATUSES, ALL_STATUSES, CLOSED_STATUSES,
    CONTACT_ACCEPTED, CONTACT_AWAITING, CONTACT_NOT_CONTACTED, CONTACT_POSTPONED, CONTACT_REJECTED,
    DEFAULT_SETTINGS, ITEM_HEADERS, LOG_HEADERS, ORDERS_HEADERS, SETTINGS_HEADERS, UNDO_HEADERS,
    STATUS_AVAILABLE, STATUS_CANCELLED, STATUS_CONTACTED, STATUS_NOT_PICKED, STATUS_PARTIAL,
    STATUS_PENDING, STATUS_PICKED_UP, STATUS_UNAVAILABLE,
    MAX_IMAGE_SIZE, ALLOWED_IMAGE_EXTS,
    add_days, now_str, today_str,
)
TZ = ZoneInfo("Asia/Riyadh")
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    order_id TEXT PRIMARY KEY,
    customer_name TEXT NOT NULL,
    phone TEXT NOT NULL,
    product_name TEXT NOT NULL DEFAULT '', quantity INTEGER NOT NULL DEFAULT 0,
    order_date TEXT NOT NULL DEFAULT '', available_date TEXT NOT NULL DEFAULT '',
    status TEXT NOT NULL, contact_status TEXT NOT NULL DEFAULT '', last_contact_date TEXT NOT NULL DEFAULT '',
    next_followup_date TEXT NOT NULL DEFAULT '', pickup_date TEXT NOT NULL DEFAULT '', notes TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL, updated_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS order_items (
    item_id TEXT PRIMARY KEY,
    order_id TEXT NOT NULL REFERENCES orders(order_id) ON DELETE CASCADE,
    product_name TEXT NOT NULL, quantity INTEGER NOT NULL DEFAULT 1,
    image_path TEXT NOT NULL DEFAULT '', availability_status TEXT NOT NULL DEFAULT 'بانتظار التوفر',
    available_price TEXT NOT NULL DEFAULT '', discounted_price TEXT NOT NULL DEFAULT '', unavailable_reason TEXT NOT NULL DEFAULT '',
    availability_note TEXT NOT NULL DEFAULT '', price_confirmation_required TEXT NOT NULL DEFAULT '', available_at TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL, customer_decision TEXT NOT NULL DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id);
CREATE TABLE IF NOT EXISTS activity_log (
    log_id TEXT PRIMARY KEY, order_id TEXT NOT NULL, action TEXT NOT NULL,
    old_status TEXT NOT NULL DEFAULT '', new_status TEXT NOT NULL DEFAULT '', note TEXT NOT NULL DEFAULT '',
    created_at TEXT NOT NULL, user_name TEXT NOT NULL DEFAULT 'موظف'
);
CREATE INDEX IF NOT EXISTS idx_activity_order_id ON activity_log(order_id);
CREATE TABLE IF NOT EXISTS undo_history (
    undo_id TEXT PRIMARY KEY, order_id TEXT NOT NULL, action TEXT NOT NULL, snapshot_json TEXT NOT NULL,
    created_at TEXT NOT NULL, undone_at TEXT NOT NULL DEFAULT '', user_name TEXT NOT NULL DEFAULT 'موظف'
);
CREATE INDEX IF NOT EXISTS idx_undo_order_id ON undo_history(order_id);
CREATE TABLE IF NOT EXISTS undo_item_images (
    undo_id TEXT NOT NULL REFERENCES undo_history(undo_id) ON DELETE CASCADE,
    item_id TEXT NOT NULL, image_path TEXT NOT NULL, filename TEXT NOT NULL,
    content_type TEXT NOT NULL, data BYTEA NOT NULL, created_at TEXT NOT NULL,
    PRIMARY KEY (undo_id, item_id, image_path)
);
CREATE INDEX IF NOT EXISTS idx_undo_item_images_undo_id ON undo_item_images(undo_id);
CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT NOT NULL DEFAULT '');
CREATE TABLE IF NOT EXISTS item_images (
    image_path TEXT PRIMARY KEY, order_id TEXT NOT NULL REFERENCES orders(order_id) ON DELETE CASCADE,
    item_id TEXT NOT NULL REFERENCES order_items(item_id) ON DELETE CASCADE, filename TEXT NOT NULL,
    content_type TEXT NOT NULL, data BYTEA NOT NULL, created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_item_images_order_item ON item_images(order_id, item_id);
CREATE TABLE IF NOT EXISTS backups (
    filename TEXT PRIMARY KEY, reason TEXT NOT NULL, created_at TEXT NOT NULL, data BYTEA NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_backups_created_at ON backups(created_at DESC);
"""
def _clean_name(value):
    value = str(value or '').strip(); return value or 'موظف'
def _safe_filename(name): return os.path.basename(str(name or '')).strip()
def _row_to_order(row):
    if not row: return None
    out=dict(row)
    return {'Order_ID':out.get('order_id',''),'Customer_Name':out.get('customer_name',''),'Phone':out.get('phone',''),'Product_Name':out.get('product_name',''),'Quantity':out.get('quantity',0) or 0,'Order_Date':out.get('order_date',''),'Available_Date':out.get('available_date',''),'Status':out.get('status',''),'Contact_Status':out.get('contact_status',''),'Last_Contact_Date':out.get('last_contact_date',''),'Next_Followup_Date':out.get('next_followup_date',''),'Pickup_Date':out.get('pickup_date',''),'Notes':out.get('notes',''),'Created_At':out.get('created_at',''),'Updated_At':out.get('updated_at','')}
def _row_to_item(row):
    return {'Item_ID':row.get('item_id',''),'Order_ID':row.get('order_id',''),'Product_Name':row.get('product_name',''),'Quantity':row.get('quantity',1),'Image_Path':row.get('image_path',''),'Availability_Status':row.get('availability_status','بانتظار التوفر'),'Available_Price':row.get('available_price',''),'Discounted_Price':row.get('discounted_price',''),'Unavailable_Reason':row.get('unavailable_reason',''),'Availability_Note':row.get('availability_note',''),'Price_Confirmation_Required':row.get('price_confirmation_required',''),'Available_At':row.get('available_at',''),'Created_At':row.get('created_at',''),'Customer_Decision':row.get('customer_decision','')}
class CloudDB:
    def __init__(self):
        self.database_url=os.environ.get('DATABASE_URL','').strip()
        if not self.database_url: raise RuntimeError('DATABASE_URL غير مضبوط. أضف رابط PostgreSQL/Supabase إلى متغيرات البيئة.')
        self.ensure_db()
    def _connect(self): return psycopg.connect(self.database_url,row_factory=dict_row,connect_timeout=10)
    def ensure_db(self):
        with self._connect() as conn:
            conn.execute(SCHEMA_SQL); conn.execute("ALTER TABLE order_items ADD COLUMN IF NOT EXISTS customer_decision TEXT NOT NULL DEFAULT ''")
            existing={r['key'] for r in conn.execute('SELECT key FROM settings').fetchall()}
            for key,value in DEFAULT_SETTINGS.items():
                if key not in existing: conn.execute('INSERT INTO settings(key,value) VALUES (%s,%s)',(key,str(value)))
    def storage_info(self):
        with self._connect() as conn:
            orders=conn.execute('SELECT COUNT(*) AS c FROM orders').fetchone()['c']; items=conn.execute('SELECT COUNT(*) AS c FROM order_items').fetchone()['c']; images=conn.execute('SELECT COUNT(*) AS c FROM item_images').fetchone()['c']; backups=conn.execute('SELECT COUNT(*) AS c FROM backups').fetchone()['c']
        return {'type':'postgresql','provider':'supabase_or_postgresql','persistent':True,'orders':int(orders),'items':int(items),'images':int(images),'backups':int(backups)}
    def _next_order_id(self,conn):
        conn.execute("SELECT pg_advisory_xact_lock(hashtext('ezz_order_id_sequence'))"); row=conn.execute("SELECT COALESCE(MAX(NULLIF(SPLIT_PART(order_id, '-', 2), '')::INT), 0) AS n FROM orders WHERE order_id ~ '^ORD-[0-9]+$'").fetchone(); return f"ORD-{int(row['n'])+1:05d}"
    def _next_item_id(self,conn):
        conn.execute("SELECT pg_advisory_xact_lock(hashtext('ezz_item_id_sequence'))"); row=conn.execute("SELECT COALESCE(MAX(NULLIF(SPLIT_PART(item_id, '-', 2), '')::INT), 0) AS n FROM order_items WHERE item_id ~ '^ITEM-[0-9]+$'").fetchone(); return f"ITEM-{int(row['n'])+1:06d}"
    def _next_log_id(self,conn): return f"LOG-{uuid.uuid4().hex[:12].upper()}"
    def _next_undo_id(self,conn): return f"UNDO-{uuid.uuid4().hex[:12].upper()}"
    def _fetch_order(self,conn,order_id): return _row_to_order(conn.execute('SELECT * FROM orders WHERE order_id=%s',(str(order_id),)).fetchone())
    def _fetch_items(self,conn,order_id): return [_row_to_item(dict(r)) for r in conn.execute('SELECT * FROM order_items WHERE order_id=%s ORDER BY created_at,item_id',(str(order_id),)).fetchall()]
    def _item_is_rejected(self,order,item):
        decision=str(item.get('Customer_Decision') or '').strip()
        if decision=='rejected': return True
        if not decision and order.get('Status')==STATUS_CANCELLED and order.get('Contact_Status')==CONTACT_REJECTED and item.get('Availability_Status')!='بانتظار التوفر': return True
        return False
    def _derive_workflow_status(self,order,items=None):
        items=items if items is not None else (order.get('Items') or []); active=[i for i in items if not self._item_is_rejected(order,i)]
        if not active:return STATUS_CANCELLED
        states=[str(i.get('Availability_Status') or 'بانتظار التوفر').strip() for i in active]
        if all(x=='بانتظار التوفر' for x in states):return STATUS_PENDING
        if all(x=='متوفر' for x in states):return STATUS_AVAILABLE
        if all(x=='غير متوفر' for x in states):return STATUS_UNAVAILABLE
        return STATUS_PARTIAL
    def _attach_items(self,orders,item_groups):
        for order in orders:
            items=item_groups.get(str(order['Order_ID']),[])
            if not items and order.get('Product_Name'):items=[{'Item_ID':'','Order_ID':order.get('Order_ID'),'Product_Name':order.get('Product_Name'),'Quantity':order.get('Quantity') or 1,'Image_Path':'','Availability_Status':'بانتظار التوفر','Available_Price':'','Discounted_Price':'','Unavailable_Reason':'','Availability_Note':'','Price_Confirmation_Required':'','Available_At':'','Created_At':order.get('Created_At',''),'Customer_Decision':''}]
            order['Items']=items
            if items:
                order['Product_Name']='، '.join(f"{i.get('Product_Name','')} × {i.get('Quantity',1)}" for i in items); order['Quantity']=sum(int(i.get('Quantity') or 0) for i in items)
            else: order['Product_Name']=''; order['Quantity']=0
        return orders
    def get_all_orders(self):
        with self._connect() as conn:
            orders=[_row_to_order(r) for r in conn.execute('SELECT * FROM orders ORDER BY created_at DESC').fetchall()]; rows=conn.execute('SELECT * FROM order_items ORDER BY created_at,item_id').fetchall()
        groups={}
        for r in rows:
            i=_row_to_item(dict(r)); groups.setdefault(str(i['Order_ID']),[]).append(i)
        return self._attach_items(orders,groups)
    def get_order(self,order_id):
        with self._connect() as conn:
            order=self._fetch_order(conn,order_id)
            if not order:return None
            order['Items']=self._fetch_items(conn,order_id)
            if order['Items']:
                order['Product_Name']='، '.join(f"{i.get('Product_Name','')} × {i.get('Quantity',1)}" for i in order['Items']); order['Quantity']=sum(int(i.get('Quantity') or 0) for i in order['Items'])
            return order
    def get_activity_log(self,order_id=None):
        with self._connect() as conn: rows=conn.execute('SELECT * FROM activity_log ORDER BY created_at DESC').fetchall() if order_id is None else conn.execute('SELECT * FROM activity_log WHERE order_id=%s ORDER BY created_at DESC',(str(order_id),)).fetchall()
        return [{'Log_ID':r['log_id'],'Order_ID':r['order_id'],'Action':r['action'],'Old_Status':r['old_status'],'New_Status':r['new_status'],'Note':r['note'],'Created_At':r['created_at'],'User':r['user_name']} for r in rows]
    def get_settings(self):
        with self._connect() as conn: rows=conn.execute('SELECT key,value FROM settings').fetchall()
        return {r['key']:r['value'] for r in rows}
    def update_settings(self,updates):
        clean={str(k):str(v) for k,v in (updates or {}).items() if str(k) in set(DEFAULT_SETTINGS)}
        with self._connect() as conn:
            for key,value in clean.items():conn.execute('INSERT INTO settings(key,value) VALUES (%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value',(key,value))
        return self.get_settings()
    def _log(self,conn,order_id,action,old_status,new_status,note,user):
        conn.execute('INSERT INTO activity_log(log_id,order_id,action,old_status,new_status,note,created_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',(self._next_log_id(conn),str(order_id),action,old_status or '',new_status or '',note or '',now_str(),_clean_name(user)))
    def _snapshot(self,conn,order_id):
        order=self._fetch_order(conn,order_id)
        return {'order':order,'items':self._fetch_items(conn,order_id)} if order else None
    def _invalidate_undo(self,conn,order_id): conn.execute("UPDATE undo_history SET undone_at=%s WHERE order_id=%s AND COALESCE(undone_at,'')=''",(now_str(),str(order_id)))
    def _add_undo(self,conn,order_id,action,snapshot,user,image_item_ids=None):
        undo_id=self._next_undo_id(conn); conn.execute('INSERT INTO undo_history(undo_id,order_id,action,snapshot_json,created_at,undone_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s)',(undo_id,str(order_id),action,json.dumps(snapshot,ensure_ascii=False),now_str(),'',_clean_name(user)))
        wanted={str(x).strip() for x in (image_item_ids or set()) if str(x).strip()}
        if not wanted:return undo_id
        ph=','.join(['%s']*len(wanted)); rows=conn.execute(f'SELECT item_id,image_path,filename,content_type,data,created_at FROM item_images WHERE order_id=%s AND item_id IN ({ph})',(str(order_id),*sorted(wanted))).fetchall()
        for row in rows:conn.execute('INSERT INTO undo_item_images(undo_id,item_id,image_path,filename,content_type,data,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)',(undo_id,str(row['item_id']),str(row['image_path']),str(row['filename']),str(row['content_type']),psycopg.Binary(bytes(row['data'])),str(row['created_at'])))
        return undo_id
    def get_undo_info(self,order_id):
        with self._connect() as conn:row=conn.execute("SELECT undo_id,action,created_at FROM undo_history WHERE order_id=%s AND COALESCE(undone_at,'')='' ORDER BY created_at DESC LIMIT 1",(str(order_id),)).fetchone()
        return {'available':False} if not row else {'available':True,'action':row['action'],'created_at':row['created_at'],'undo_id':row['undo_id']}
    def create_order(self,customer_name,phone,products,notes='',order_date=None,user='موظف'):
        clean_products=[]
        for p in products:
            try:qty=int(p.get('quantity',0))
            except (TypeError,ValueError):qty=0
            name=str(p.get('product_name','')).strip()
            if name and qty>0:clean_products.append({'product_name':name,'quantity':qty})
        if not clean_products:raise ValueError('يجب إضافة منتج واحد على الأقل')
        with self._connect() as conn:
            oid=self._next_order_id(conn);ts=now_str();odate=str(order_date or today_str());summary='، '.join(f"{p['product_name']} × {p['quantity']}" for p in clean_products);qty=sum(p['quantity'] for p in clean_products)
            conn.execute('INSERT INTO orders(order_id,customer_name,phone,product_name,quantity,order_date,available_date,status,contact_status,last_contact_date,next_followup_date,pickup_date,notes,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(oid,customer_name,phone,summary,qty,odate,'',STATUS_PENDING,CONTACT_NOT_CONTACTED,'','','',notes or '',ts,ts))
            for p in clean_products:conn.execute('INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(self._next_item_id(conn),oid,p['product_name'],p['quantity'],'','بانتظار التوفر','','','','','','',ts))
            self._log(conn,oid,'إنشاء الطلب','',STATUS_PENDING,f'تم تسجيل {len(clean_products)} منتج في الطلب',user)
        return self.get_order(oid)
    def update_order(self,order_id,fields,products=None,user='موظف',deleted_item_ids=None):
        with self._connect() as conn:
            current=self._fetch_order(conn,order_id)
            if not current:return None
            snapshot=self._snapshot(conn,order_id);self._invalidate_undo(conn,order_id);updates=dict(fields or {})
            if 'Status' in updates:raise ValueError('تغيير حالة الطلب مباشرة غير مسموح. استخدم إجراء الحالة المناسب حتى يتم التحقق من قواعد سير الطلب.')
            if 'Contact_Status' in updates:raise ValueError('تغيير حالة التواصل مباشرة غير مسموح. استخدم إجراء حالة التواصل المناسب.')
            if products is not None:
                existing_items=self._fetch_items(conn,order_id);existing_by_id={str(i['Item_ID']):i for i in existing_items};deleted_ids={str(x).strip() for x in (deleted_item_ids or []) if str(x).strip()};overlap=deleted_ids.intersection(existing_by_id);unknown_deleted=deleted_ids.difference(existing_by_id)
                if unknown_deleted:raise ValueError(f"لا يمكن حذف منتجات غير موجودة في الطلب: {', '.join(sorted(unknown_deleted))}")
                resolved=[];seen_ids=set();new_count=0
                for index,p in enumerate(products,1):
                    p=p or {};name=str(p.get('product_name') or '').strip()
                    try:qty=int(p.get('quantity',0))
                    except (TypeError,ValueError):qty=0
                    if not name:raise ValueError(f'اسم المنتج رقم {index} مطلوب')
                    if qty<=0:raise ValueError(f'كمية المنتج رقم {index} يجب أن تكون أكبر من صفر')
                    item_id=str(p.get('item_id') or p.get('Item_ID') or '').strip()
                    if item_id:
                        if item_id not in existing_by_id:raise ValueError(f'Item_ID غير موجود في هذا الطلب: {item_id}')
                        if item_id in seen_ids:raise ValueError(f'تم تكرار Item_ID في المنتجات: {item_id}')
                        if item_id in overlap:raise ValueError(f'لا يمكن تعديل وحذف نفس المنتج: {item_id}')
                        seen_ids.add(item_id);resolved.append({'item_id':item_id,'product_name':name,'quantity':qty,'is_new':False})
                    else:new_count+=1;resolved.append({'item_id':None,'product_name':name,'quantity':qty,'is_new':True})
                if len(existing_items)-len(overlap)+new_count<=0:raise ValueError('لا يمكن حفظ الطلب بدون أي منتج. أضف منتجًا أو ألغِ حذف المنتج.')
                self._add_undo(conn,order_id,'تعديل بيانات الطلب',snapshot,user,image_item_ids=overlap);ts=now_str()
                for item in resolved:
                    if item['is_new']:conn.execute('INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(self._next_item_id(conn),str(order_id),item['product_name'],item['quantity'],'','بانتظار التوفر','','','','','','',ts))
                    else:conn.execute('UPDATE order_items SET product_name=%s,quantity=%s WHERE item_id=%s AND order_id=%s',(item['product_name'],item['quantity'],item['item_id'],str(order_id)))
                for item_id in sorted(overlap):conn.execute('DELETE FROM order_items WHERE item_id=%s AND order_id=%s',(item_id,str(order_id)))
                fresh_items=self._fetch_items(conn,order_id);updates['Product_Name']='، '.join(f"{i['Product_Name']} × {i['Quantity']}" for i in fresh_items);updates['Quantity']=sum(int(i.get('Quantity') or 0) for i in fresh_items)
            else:self._add_undo(conn,order_id,'تعديل بيانات الطلب',snapshot,user)
            col_map={'Customer_Name':'customer_name','Phone':'phone','Notes':'notes','Order_Date':'order_date','Status':'status','Available_Date':'available_date','Contact_Status':'contact_status','Last_Contact_Date':'last_contact_date','Next_Followup_Date':'next_followup_date','Pickup_Date':'pickup_date','Product_Name':'product_name','Quantity':'quantity'};sets=[];params=[]
            for key,val in updates.items():
                if key in col_map:sets.append(f'{col_map[key]}=%s');params.append(val)
            if sets:sets.append('updated_at=%s');params.append(now_str());params.append(str(order_id));conn.execute(f"UPDATE orders SET {', '.join(sets)} WHERE order_id=%s",params)
            self._log(conn,order_id,'تعديل بيانات الطلب',current['Status'],updates.get('Status',current['Status']),'تم تعديل بيانات الطلب',user);return self._refresh_order_in_conn(conn,order_id)
    def _refresh_order_in_conn(self,conn,order_id):
        order=self._fetch_order(conn,order_id)
        if not order:return None
        order['Items']=self._fetch_items(conn,order_id)
        if order['Items']:order['Product_Name']='، '.join(f"{i.get('Product_Name','')} × {i.get('Quantity',1)}" for i in order['Items']);order['Quantity']=sum(int(i.get('Quantity') or 0) for i in order['Items'])
        return order
    def undo_last(self,order_id,user='موظف'):
        with self._connect() as conn:
            row=conn.execute("SELECT * FROM undo_history WHERE order_id=%s AND COALESCE(undone_at,'')='' ORDER BY created_at DESC LIMIT 1",(str(order_id),)).fetchone();current=self._fetch_order(conn,order_id)
            if not row:return {'error':'لا يوجد إجراء يمكن التراجع عنه لهذا الطلب','code':409}
            if not current:return {'error':'الطلب غير موجود','code':404}
            snapshot=json.loads(row['snapshot_json']);order_data=snapshot.get('order',{});image_rows=conn.execute('SELECT item_id,image_path,filename,content_type,data,created_at FROM undo_item_images WHERE undo_id=%s ORDER BY item_id,image_path',(row['undo_id'],)).fetchall();images_by_item={}
            for image in image_rows:images_by_item.setdefault(str(image['item_id']),[]).append(image)
            sets=[];params=[]
            for key,dbkey in [('Customer_Name','customer_name'),('Phone','phone'),('Product_Name','product_name'),('Quantity','quantity'),('Order_Date','order_date'),('Available_Date','available_date'),('Status','status'),('Contact_Status','contact_status'),('Last_Contact_Date','last_contact_date'),('Next_Followup_Date','next_followup_date'),('Pickup_Date','pickup_date'),('Notes','notes'),('Created_At','created_at')]:sets.append(f'{dbkey}=%s');params.append(order_data.get(key,''))
            sets.append('updated_at=%s');params.append(now_str());params.append(str(order_id));conn.execute(f"UPDATE orders SET {', '.join(sets)} WHERE order_id=%s",params)
            snapshot_items=snapshot.get('items',[]) or [];current_items=self._fetch_items(conn,order_id);current_by_id={str(i.get('Item_ID')):i for i in current_items if i.get('Item_ID')};snapshot_by_id={str(i.get('Item_ID')):i for i in snapshot_items if i.get('Item_ID')}
            for item_id in sorted(set(current_by_id)-set(snapshot_by_id)):conn.execute('DELETE FROM order_items WHERE item_id=%s AND order_id=%s',(item_id,str(order_id)))
            fields=('Product_Name','Quantity','Image_Path','Availability_Status','Available_Price','Discounted_Price','Unavailable_Reason','Availability_Note','Price_Confirmation_Required','Available_At','Created_At','Customer_Decision')
            for item_id,item in snapshot_by_id.items():
                values=tuple(item.get(k,'') for k in fields)
                if item_id in current_by_id:
                    conn.execute('UPDATE order_items SET product_name=%s,quantity=%s,image_path=%s,availability_status=%s,available_price=%s,discounted_price=%s,unavailable_reason=%s,availability_note=%s,price_confirmation_required=%s,available_at=%s,created_at=%s,customer_decision=%s WHERE item_id=%s AND order_id=%s',values+(item_id,str(order_id)))
                    expected=str(item.get('Image_Path') or '');actual=str(current_by_id[item_id].get('Image_Path') or '')
                    if expected!=actual and expected and item_id not in images_by_item:raise ValueError(f'لا يمكن التراجع بأمان عن الصورة المرتبطة بالمنتج {item_id}: بيانات الصورة السابقة غير محفوظة')
                    if not expected and actual:conn.execute('DELETE FROM item_images WHERE order_id=%s AND item_id=%s',(str(order_id),item_id))
                else:
                    conn.execute('INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at,customer_decision) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(item_id,str(order_id),item.get('Product_Name',''),item.get('Quantity',1),item.get('Image_Path',''),item.get('Availability_Status','بانتظار التوفر'),item.get('Available_Price',''),item.get('Discounted_Price',''),item.get('Unavailable_Reason',''),item.get('Availability_Note',''),item.get('Price_Confirmation_Required',''),item.get('Available_At',''),item.get('Created_At') or now_str(),item.get('Customer_Decision','')))
                    if item.get('Image_Path') and item_id not in images_by_item:raise ValueError(f'لا يمكن التراجع بأمان عن الصورة المرتبطة بالمنتج {item_id}: بيانات الصورة السابقة غير محفوظة')
            for item_id,images in images_by_item.items():
                if item_id not in snapshot_by_id:raise ValueError(f'صورة Undo مرتبطة بمنتج غير موجود في snapshot: {item_id}')
                conn.execute('DELETE FROM item_images WHERE order_id=%s AND item_id=%s',(str(order_id),item_id))
                for image in images:conn.execute('INSERT INTO item_images(image_path,order_id,item_id,filename,content_type,data,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)',(str(image['image_path']),str(order_id),item_id,str(image['filename']),str(image['content_type']),psycopg.Binary(bytes(image['data'])),str(image['created_at'])))
            conn.execute('UPDATE undo_history SET undone_at=%s WHERE undo_id=%s',(now_str(),row['undo_id']));self._log(conn,order_id,f"تراجع عن: {row['action']}",current['Status'],order_data.get('Status',''),'تم التراجع عن آخر تغيير للمستخدم',user);return {'order':self._refresh_order_in_conn(conn,order_id),'undone_action':row['action']}
    def reset_all_data(self,user='موظف'):
        self.create_manual_backup(reason='auto')
        with self._connect() as conn:conn.execute('TRUNCATE activity_log, undo_item_images, undo_history, item_images, order_items, orders')
    def _workbook_bytes(self):
        wb=Workbook();ws=wb.active;ws.title='Orders';ws.append(ORDERS_HEADERS);wi=wb.create_sheet('Order_Items');wi.append(ITEM_HEADERS);wl=wb.create_sheet('Activity_Log');wl.append(LOG_HEADERS);wu=wb.create_sheet('Undo_History');wu.append(UNDO_HEADERS);wsset=wb.create_sheet('Settings');wsset.append(SETTINGS_HEADERS)
        with self._connect() as conn:order_rows=conn.execute('SELECT * FROM orders ORDER BY created_at').fetchall();item_rows=conn.execute('SELECT * FROM order_items ORDER BY created_at,item_id').fetchall();log_rows=conn.execute('SELECT * FROM activity_log ORDER BY created_at').fetchall();undo_rows=conn.execute('SELECT * FROM undo_history ORDER BY created_at').fetchall();settings=conn.execute('SELECT key,value FROM settings ORDER BY key').fetchall()
        for r in order_rows:d=_row_to_order(r);ws.append([d[h] for h in ORDERS_HEADERS])
        for r in item_rows:d=_row_to_item(dict(r));wi.append([d[h] for h in ITEM_HEADERS])
        for r in log_rows:wl.append([r['log_id'],r['order_id'],r['action'],r['old_status'],r['new_status'],r['note'],r['created_at'],r['user_name']])
        for r in undo_rows:wu.append([r['undo_id'],r['order_id'],r['action'],r['snapshot_json'],r['created_at'],r['undone_at'],r['user_name']])
        for r in settings:wsset.append([r['key'],r['value']])
        bio=io.BytesIO();wb.save(bio);wb.close();return bio.getvalue()
    def _backup_bytes(self,reason):
        xlsx=self._workbook_bytes();bio=io.BytesIO()
        with zipfile.ZipFile(bio,'w',zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('data/pharmacy_orders.xlsx',xlsx)
            with self._connect() as conn:rows=conn.execute('SELECT image_path,data FROM item_images ORDER BY image_path').fetchall()
            for r in rows:zf.writestr('uploads/'+r['image_path'],bytes(r['data']))
            zf.writestr('backup_info.txt',f'Ezz Pharmacy cloud backup\nCreated: {now_str()}\nReason: {reason}\n')
        return bio.getvalue()
    def create_manual_backup(self,reason='manual'):
        reason='manual' if reason=='manual' else 'auto';filename=f"backup_{reason}_{datetime.now(TZ).strftime('%Y-%m-%d_%H%M%S_%f')}.zip";data=self._backup_bytes(reason)
        with self._connect() as conn:conn.execute('INSERT INTO backups(filename,reason,created_at,data) VALUES (%s,%s,%s,%s)',(filename,reason,now_str(),psycopg.Binary(data)));return filename
    def list_backups(self):
        with self._connect() as conn:rows=conn.execute('SELECT filename,reason,created_at,octet_length(data) AS bytes FROM backups ORDER BY created_at DESC').fetchall()
        return [{'filename':r['filename'],'size_kb':round(int(r['bytes'])/1024,1),'created_at':r['created_at'],'reason':'تلقائية' if r['reason']=='auto' else 'يدوية'} for r in rows]
    def restore_backup(self,filename):
        safe=_safe_filename(filename)
        with self._connect() as conn:row=conn.execute('SELECT data FROM backups WHERE filename=%s',(safe,)).fetchone()
        if not row:return False
        self.create_manual_backup(reason='auto');self._restore_from_zip(bytes(row['data']));return True
    def _validate_zip_names(self,zf):
        for info in zf.infolist():
            name=info.filename.replace('\\','/')
            if name.startswith('/') or '..' in name.split('/'):raise ValueError('ملف النسخة الاحتياطية يحتوي مسارًا غير آمن')
    def _read_sheet_dicts(self,ws):
        headers=[str(c.value or '') for c in ws[1]];out=[]
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not row or all(v is None or v=='' for v in row):continue
            out.append({headers[i]:(row[i] if i<len(row) and row[i] is not None else '') for i in range(len(headers))})
        return out
    def _restore_from_zip(self,blob):
        with zipfile.ZipFile(io.BytesIO(blob),'r') as zf:
            self._validate_zip_names(zf);names={n.replace('\\','/'):n for n in zf.namelist()};xlsx_name='data/pharmacy_orders.xlsx' if 'data/pharmacy_orders.xlsx' in names else 'pharmacy_orders.xlsx'
            if xlsx_name not in names:raise ValueError('النسخة الاحتياطية لا تحتوي على ملف البيانات')
            xlsx=zf.read(names[xlsx_name]);images={k[8:]:zf.read(v) for k,v in names.items() if k.startswith('uploads/') and not k.endswith('/')}
        self._replace_from_xlsx(xlsx,images)
    def _replace_from_xlsx(self,xlsx_bytes,images=None):
        wb=load_workbook(io.BytesIO(xlsx_bytes),read_only=True,data_only=False)
        if 'Orders' not in wb.sheetnames:wb.close();raise ValueError('ملف البيانات لا يحتوي ورقة Orders المطلوبة')
        orders=self._read_sheet_dicts(wb['Orders']);
        if not any(any(k in r for k in ('Order_ID','Customer_Name','Phone')) for r in orders):wb.close();raise ValueError('ملف البيانات لا يبدو كملف طلبات صالح')
        items=self._read_sheet_dicts(wb['Order_Items']) if 'Order_Items' in wb.sheetnames else [];logs=self._read_sheet_dicts(wb['Activity_Log']) if 'Activity_Log' in wb.sheetnames else [];undos=self._read_sheet_dicts(wb['Undo_History']) if 'Undo_History' in wb.sheetnames else [];settings=self._read_sheet_dicts(wb['Settings']) if 'Settings' in wb.sheetnames else [];wb.close()
        with self._connect() as conn:
            conn.execute('TRUNCATE activity_log, undo_item_images, undo_history, item_images, order_items, orders')
            for r in orders:
                oid=str(r.get('Order_ID') or '').strip()
                if not oid:continue
                status=r.get('Status') or STATUS_PENDING;contact=r.get('Contact_Status') or (CONTACT_AWAITING if status in (STATUS_CONTACTED,STATUS_NOT_PICKED) else CONTACT_NOT_CONTACTED)
                conn.execute('INSERT INTO orders(order_id,customer_name,phone,product_name,quantity,order_date,available_date,status,contact_status,last_contact_date,next_followup_date,pickup_date,notes,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(oid,str(r.get('Customer_Name') or ''),str(r.get('Phone') or ''),str(r.get('Product_Name') or ''),int(r.get('Quantity') or 0),str(r.get('Order_Date') or ''),str(r.get('Available_Date') or ''),status,contact,str(r.get('Last_Contact_Date') or ''),str(r.get('Next_Followup_Date') or ''),str(r.get('Pickup_Date') or ''),str(r.get('Notes') or ''),str(r.get('Created_At') or now_str()),str(r.get('Updated_At') or now_str())))
            for r in items:
                iid=str(r.get('Item_ID') or '').strip();oid=str(r.get('Order_ID') or '').strip()
                if not iid or not oid:continue
                conn.execute('INSERT INTO order_items(item_id,order_id,product_name,quantity,image_path,availability_status,available_price,discounted_price,unavailable_reason,availability_note,price_confirmation_required,available_at,created_at,customer_decision) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)',(iid,oid,str(r.get('Product_Name') or ''),int(r.get('Quantity') or 1),str(r.get('Image_Path') or ''),str(r.get('Availability_Status') or 'بانتظار التوفر'),str(r.get('Available_Price') or ''),str(r.get('Discounted_Price') or ''),str(r.get('Unavailable_Reason') or ''),str(r.get('Availability_Note') or ''),str(r.get('Price_Confirmation_Required') or ''),str(r.get('Available_At') or ''),str(r.get('Created_At') or now_str()),str(r.get('Customer_Decision') or '')))
            for r in logs:conn.execute('INSERT INTO activity_log(log_id,order_id,action,old_status,new_status,note,created_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)',(str(r.get('Log_ID') or uuid.uuid4().hex),str(r.get('Order_ID') or ''),str(r.get('Action') or ''),str(r.get('Old_Status') or ''),str(r.get('New_Status') or ''),str(r.get('Note') or ''),str(r.get('Created_At') or now_str()),str(r.get('User') or 'موظف')))
            for r in undos:conn.execute('INSERT INTO undo_history(undo_id,order_id,action,snapshot_json,created_at,undone_at,user_name) VALUES (%s,%s,%s,%s,%s,%s,%s)',(str(r.get('Undo_ID') or uuid.uuid4().hex),str(r.get('Order_ID') or ''),str(r.get('Action') or ''),str(r.get('Snapshot_JSON') or '{}'),str(r.get('Created_At') or now_str()),str(r.get('Undone_At') or ''),str(r.get('User') or 'موظف')))
            for r in settings:
                k=str(r.get('Key') or '').strip()
                if k:conn.execute('INSERT INTO settings(key,value) VALUES (%s,%s) ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value',(k,str(r.get('Value') or '')))
            if images:
                for path,data in images.items():
                    path=path.replace('\\','/').lstrip('/');m=re.match(r'^([^/]+)/([^/]+)',path)
                    if not m:continue
                    oid=m.group(1);filename=os.path.basename(path);iid=filename.split('_',1)[0];item=conn.execute('SELECT item_id FROM order_items WHERE item_id=%s AND order_id=%s',(iid,oid)).fetchone()
                    if not item:continue
                    ext=os.path.splitext(filename)[1].lower();ctype={'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','webp':'image/webp'}.get(ext.lstrip('.'),'application/octet-stream')
                    conn.execute('INSERT INTO item_images(image_path,order_id,item_id,filename,content_type,data,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT(image_path) DO UPDATE SET data=EXCLUDED.data,content_type=EXCLUDED.content_type,filename=EXCLUDED.filename',(path,oid,iid,filename,ctype,psycopg.Binary(data),now_str()))
__all__=['CloudDB']
