# -*- coding: utf-8 -*-
"""Read-only Excel export for the current application data."""
import os
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

from flask import jsonify, send_file
from openpyxl import Workbook


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    if headers:
        ws.auto_filter.ref = ws.dimensions


def _build_workbook(db):
    path = getattr(db, "DB_PATH", "")
    # Current production mode is Excel-backed. Send the real data file directly so
    # the browser downloads the exact workbook the application is using.
    if path and os.path.isfile(path):
        return path, None, None

    # Cloud/PostgreSQL mode: construct an Excel snapshot from the persistent DB.
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    orders = db.get_all_orders() or []
    order_headers = ["Order_ID","Customer_Name","Phone","Product_Name","Quantity","Order_Date","Available_Date","Status","Contact_Status","Last_Contact_Date","Next_Followup_Date","Pickup_Date","Notes","Created_At","Updated_At"]
    _write_sheet(ws, order_headers, orders)

    items = []
    for order in orders:
        for item in order.get("Items", []) or []:
            row = dict(item)
            row.setdefault("Order_ID", order.get("Order_ID", ""))
            items.append(row)
    _write_sheet(wb.create_sheet("Order_Items"), ["Item_ID","Order_ID","Product_Name","Quantity","Image_Path","Availability_Status","Available_Price","Discounted_Price","Unavailable_Reason","Availability_Note","Price_Confirmation_Required","Available_At","Created_At"], items)

    try:
        logs = db.get_activity_log() or []
    except Exception:
        logs = []
    _write_sheet(wb.create_sheet("Activity_Log"), ["Log_ID","Order_ID","Action","Old_Status","New_Status","Note","Created_At","User"], logs)

    try:
        settings = db.get_settings() or {}
        settings_rows = [{"Key": k, "Value": v} for k,v in settings.items()]
    except Exception:
        settings_rows = []
    _write_sheet(wb.create_sheet("Settings"), ["Key","Value"], settings_rows)

    out = BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return out, len(orders), len(items)


def install_data_export(app, db):
    if getattr(app, "_ezz_data_export_installed", False):
        return
    app._ezz_data_export_installed = True
    auth = app.extensions.get("ezz_auth")
    current_user = auth.get("current_user") if auth else (lambda: None)

    @app.get("/api/data/export-xlsx")
    def export_current_xlsx():
        if not current_user():
            return jsonify({"error": "تسجيل الدخول مطلوب", "authenticated": False}), 401
        try:
            exported, order_count, item_count = _build_workbook(db)
            ts = datetime.now(ZoneInfo("Asia/Riyadh")).strftime("%Y-%m-%d_%H%M%S")
            filename = f"Ezz_Pharmacy_Backup_{ts}.xlsx"
            if isinstance(exported, str):
                response = send_file(exported, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename, max_age=0, conditional=False)
            else:
                response = send_file(exported, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename, max_age=0, conditional=False)
            response.headers["Cache-Control"] = "no-store"
            if order_count is not None: response.headers["X-Export-Orders"] = str(order_count)
            if item_count is not None: response.headers["X-Export-Items"] = str(item_count)
            return response
        except Exception as exc:
            app.logger.exception("Excel export failed")
            return jsonify({"error": f"تعذر تصدير البيانات: {exc}"}), 500
